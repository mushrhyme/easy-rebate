"""
PDF를 엑셀로 변환하는 유틸리티 모듈

PDF의 테이블을 추출하여 엑셀 파일로 변환하거나, 
DataFrame으로 변환하여 LLM에 전달할 수 있습니다.
"""

from pathlib import Path
from typing import Optional, List, Dict, Any
import sys
import pandas as pd
import fitz  # PyMuPDF

# xlrd가 Python 2 문법을 사용하여 Python 3에서 오류 발생 방지
# pandas가 xlrd를 import하지 않도록 sys.modules에서 제거
def _prevent_xlrd_import():
    """xlrd import를 방지하여 Python 2 문법 오류를 예방"""
    if 'xlrd' in sys.modules:
        del sys.modules['xlrd']
    # pandas.io.excel 모듈에서도 xlrd 참조 제거 시도
    try:
        import pandas.io.excel._xlrd
        if hasattr(pandas.io.excel._xlrd, 'xlrd'):
            delattr(pandas.io.excel._xlrd, 'xlrd')
    except (ImportError, AttributeError):
        pass


class PdfToExcelConverter:
    """
    PDF를 엑셀로 변환하는 클래스
    
    여러 방법을 지원:
    1. tabula-py: 테이블 추출 (Java 필요)
    2. pdfplumber: 테이블 추출 (더 정확)
    3. PyMuPDF + pandas: 텍스트를 파싱하여 테이블로 변환
    """
    
    def __init__(self, method: str = "pdfplumber"):
        """
        Args:
            method: 변환 방법 ("tabula", "pdfplumber", "pymupdf")
        """
        self.method = method
        self._check_dependencies()
    
    def _check_dependencies(self):
        """필요한 라이브러리 확인"""
        if self.method == "tabula":
            try:
                import tabula
            except ImportError:
                raise ImportError(
                    "tabula-py가 설치되지 않았습니다.\n"
                    "pip install tabula-py로 설치하세요.\n"
                    "참고: Java가 필요합니다."
                )
        elif self.method == "pdfplumber":
            try:
                import pdfplumber
            except ImportError:
                raise ImportError(
                    "pdfplumber가 설치되지 않았습니다.\n"
                    "pip install pdfplumber로 설치하세요."
                )
    
    def extract_tables_with_tabula(
        self, 
        pdf_path: Path, 
        page_num: int
    ) -> List[pd.DataFrame]:
        """
        tabula-py를 사용하여 PDF에서 테이블 추출
        
        Args:
            pdf_path: PDF 파일 경로
            page_num: 페이지 번호 (1부터 시작)
            
        Returns:
            추출된 테이블 리스트 (DataFrame)
        """
        import tabula
        
        try:
            # tabula는 페이지 번호를 1부터 시작
            tables = tabula.read_pdf(
                str(pdf_path),
                pages=page_num,
                multiple_tables=True,
                pandas_options={'header': None}  # 헤더 자동 감지 비활성화
            )
            return tables if tables else []
        except Exception as e:
            print(f"⚠️ tabula 테이블 추출 실패 ({pdf_path}, 페이지 {page_num}): {e}")
            return []
    
    def extract_tables_with_pdfplumber(
        self, 
        pdf_path: Path, 
        page_num: int
    ) -> List[pd.DataFrame]:
        """
        pdfplumber를 사용하여 PDF에서 테이블 추출
        
        Args:
            pdf_path: PDF 파일 경로
            page_num: 페이지 번호 (1부터 시작)
            
        Returns:
            추출된 테이블 리스트 (DataFrame)
        """
        import pdfplumber
        
        try:
            tables = []
            with pdfplumber.open(pdf_path) as pdf:
                if page_num < 1 or page_num > len(pdf.pages):
                    return []
                
                page = pdf.pages[page_num - 1]
                extracted_tables = page.extract_tables()
                
                for table in extracted_tables:
                    if table:
                        # 테이블을 DataFrame으로 변환
                        df = pd.DataFrame(table[1:], columns=table[0] if table else None)
                        tables.append(df)
            
            return tables
        except Exception as e:
            print(f"⚠️ pdfplumber 테이블 추출 실패 ({pdf_path}, 페이지 {page_num}): {e}")
            return []
    
    def extract_text_as_table_with_pymupdf(
        self, 
        pdf_path: Path, 
        page_num: int
    ) -> Optional[pd.DataFrame]:
        """
        PyMuPDF로 텍스트를 추출하고 간단한 파싱으로 테이블 형태로 변환
        
        Args:
            pdf_path: PDF 파일 경로
            page_num: 페이지 번호 (1부터 시작)
            
        Returns:
            DataFrame (테이블 형태로 파싱된 텍스트)
        """
        try:
            doc = fitz.open(pdf_path)
            if page_num < 1 or page_num > doc.page_count:
                return None
            
            page = doc.load_page(page_num - 1)
            text = page.get_text()
            doc.close()
            
            if not text:
                return None
            
            # 텍스트를 줄 단위로 분리
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            
            # 간단한 테이블 파싱 (탭 또는 공백으로 구분된 데이터)
            rows = []
            for line in lines:
                # 탭 또는 여러 공백으로 분리
                parts = [p.strip() for p in line.split('\t') if p.strip()]
                if not parts:
                    # 탭이 없으면 공백으로 분리 시도
                    parts = [p.strip() for p in line.split() if p.strip()]
                
                if parts:
                    rows.append(parts)
            
            if not rows:
                return None
            
            # 모든 행의 최대 길이 구하기
            max_cols = max(len(row) for row in rows) if rows else 0
            
            # 모든 행을 같은 길이로 맞추기
            normalized_rows = []
            for row in rows:
                normalized_row = row + [''] * (max_cols - len(row))
                normalized_rows.append(normalized_row)
            
            # DataFrame 생성
            df = pd.DataFrame(normalized_rows)
            
            # 첫 번째 행을 헤더로 사용 (선택적)
            if len(df) > 1:
                # 첫 번째 행이 헤더처럼 보이면 사용
                df.columns = [f"Column_{i+1}" for i in range(max_cols)]
            
            return df
            
        except Exception as e:
            print(f"⚠️ PyMuPDF 테이블 추출 실패 ({pdf_path}, 페이지 {page_num}): {e}")
            return None
    
    def extract_tables(
        self, 
        pdf_path: Path, 
        page_num: int
    ) -> List[pd.DataFrame]:
        """
        지정된 방법으로 PDF에서 테이블 추출
        
        Args:
            pdf_path: PDF 파일 경로
            page_num: 페이지 번호 (1부터 시작)
            
        Returns:
            추출된 테이블 리스트 (DataFrame)
        """
        if self.method == "tabula":
            return self.extract_tables_with_tabula(pdf_path, page_num)
        elif self.method == "pdfplumber":
            return self.extract_tables_with_pdfplumber(pdf_path, page_num)
        else:  # pymupdf
            df = self.extract_text_as_table_with_pymupdf(pdf_path, page_num)
            return [df] if df is not None else []
    
    def convert_to_excel(
        self, 
        pdf_path: Path, 
        page_num: int,
        output_path: Optional[Path] = None
    ) -> Optional[Path]:
        """
        PDF 페이지를 엑셀 파일로 변환
        
        Args:
            pdf_path: PDF 파일 경로
            page_num: 페이지 번호 (1부터 시작)
            output_path: 출력 엑셀 파일 경로 (None이면 자동 생성)
            
        Returns:
            생성된 엑셀 파일 경로 또는 None
        """
        tables = self.extract_tables(pdf_path, page_num)
        
        if not tables:
            print(f"⚠️ 테이블을 찾을 수 없습니다 ({pdf_path}, 페이지 {page_num})")
            return None
        
        # 출력 경로 자동 생성
        if output_path is None:
            pdf_name = pdf_path.stem
            output_path = pdf_path.parent / f"{pdf_name}_Page{page_num}.xlsx"
        
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for idx, df in enumerate(tables):
                    sheet_name = f"Table_{idx+1}" if len(tables) > 1 else "Sheet1"
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            print(f"✅ 엑셀 파일 생성 완료: {output_path}")
            return output_path
        except Exception as e:
            print(f"⚠️ 엑셀 파일 생성 실패: {e}")
            return None
    
    def convert_to_text_for_llm(
        self, 
        pdf_path: Path, 
        page_num: int,
        temp_dir: Optional[Path] = None,
        keep_excel_file: bool = False
    ) -> str:
        """
        PDF를 실제 엑셀 파일(.xlsx)로 변환한 후, 그 파일을 읽어서 텍스트로 변환
        
        프로세스:
        1. PDF → 엑셀 파일(.xlsx) 생성 (실제 파일로 저장)
        2. 생성된 엑셀 파일을 pandas로 읽기
        3. 읽은 DataFrame을 텍스트로 변환
        4. 그 텍스트를 반환 (LLM에 전달)
        
        Args:
            pdf_path: PDF 파일 경로
            page_num: 페이지 번호 (1부터 시작)
            temp_dir: 엑셀 파일 저장 디렉토리 (None이면 pdf_path의 부모 디렉토리)
            keep_excel_file: 엑셀 파일을 유지할지 여부 (False면 처리 후 삭제)
            
        Returns:
            LLM에 전달할 수 있는 텍스트 형식 (엑셀 파일에서 읽은 내용)
        """
        # 저장 디렉토리 설정
        if temp_dir is None:
            temp_dir = pdf_path.parent
        else:
            temp_dir = Path(temp_dir)
            temp_dir.mkdir(parents=True, exist_ok=True)
        
        # 엑셀 파일 경로 생성
        pdf_name = pdf_path.stem
        excel_path = temp_dir / f"{pdf_name}_Page{page_num}.xlsx"
        
        try:
            # 1단계: PDF를 실제 엑셀 파일(.xlsx)로 변환
            created_excel_path = self.convert_to_excel(pdf_path, page_num, output_path=excel_path)
            
            if not created_excel_path or not created_excel_path.exists():
                # 엑셀 변환 실패 시 일반 텍스트 추출
                print(f"⚠️ 엑셀 변환 실패, PyMuPDF 텍스트로 폴백 ({pdf_path}, 페이지 {page_num})")
                try:
                    doc = fitz.open(pdf_path)
                    page = doc.load_page(page_num - 1)
                    text = page.get_text()
                    doc.close()
                    return text.strip() if text else ""
                except:
                    return ""
            
            # 2단계: 생성된 엑셀 파일을 읽기
            # xlrd import 방지 (Python 2 문법 오류 예방)
            _prevent_xlrd_import()
            try:
                excel_df = pd.read_excel(created_excel_path, sheet_name=None, engine='openpyxl')  # 모든 시트 읽기
            except (SyntaxError, ImportError) as e:
                # xlrd 관련 오류 발생 시 openpyxl을 직접 사용
                if 'xlrd' in str(e).lower() or 'print' in str(e).lower():
                    print(f"⚠️ xlrd 오류 감지, openpyxl 직접 사용: {e}")
                    from openpyxl import load_workbook
                    wb = load_workbook(created_excel_path, data_only=True)
                    excel_df = {}
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        data = []
                        for row in sheet.iter_rows(values_only=True):
                            if any(cell is not None for cell in row):  # 빈 행 제외
                                data.append(row)
                        if data:
                            # 첫 번째 행을 헤더로 사용
                            if len(data) > 1:
                                df = pd.DataFrame(data[1:], columns=data[0])
                            else:
                                df = pd.DataFrame(data)
                            excel_df[sheet_name] = df
                    wb.close()
                else:
                    raise
            
            # 3단계: 읽은 DataFrame을 텍스트로 변환
            text_parts = []
            for sheet_name, df in excel_df.items():
                if len(excel_df) > 1:
                    text_parts.append(f"=== 시트: {sheet_name} ===\n")
                else:
                    text_parts.append("=== 엑셀 데이터 ===\n")
                
                # DataFrame을 텍스트로 변환 (인덱스 없이)
                text_parts.append(df.to_string(index=False))
                text_parts.append("\n")
            
            result_text = "\n".join(text_parts)
            
            # 4단계: 엑셀 파일 삭제 (keep_excel_file=False인 경우)
            if not keep_excel_file:
                try:
                    if created_excel_path.exists():
                        created_excel_path.unlink()
                        print(f"✅ 임시 엑셀 파일 삭제: {created_excel_path}")
                except Exception as del_err:
                    print(f"⚠️ 엑셀 파일 삭제 실패 (무시): {del_err}")
            else:
                print(f"✅ 엑셀 파일 유지: {created_excel_path}")
            
            return result_text
            
        except Exception as e:
            print(f"⚠️ 엑셀 변환 및 읽기 실패 ({pdf_path}, 페이지 {page_num}): {e}")
            import traceback
            traceback.print_exc()
            
            # 실패 시 일반 텍스트 추출
            try:
                doc = fitz.open(pdf_path)
                page = doc.load_page(page_num - 1)
                text = page.get_text()
                doc.close()
                return text.strip() if text else ""
            except:
                return ""


def convert_pdf_page_to_excel(
    pdf_path: Path,
    page_num: int,
    method: str = "pdfplumber",
    output_path: Optional[Path] = None
) -> Optional[Path]:
    """
    PDF 페이지를 엑셀 파일로 변환 (간편 함수)
    
    Args:
        pdf_path: PDF 파일 경로
        page_num: 페이지 번호 (1부터 시작)
        method: 변환 방법 ("tabula", "pdfplumber", "pymupdf")
        output_path: 출력 엑셀 파일 경로 (None이면 자동 생성)
        
    Returns:
        생성된 엑셀 파일 경로 또는 None
    """
    converter = PdfToExcelConverter(method=method)
    return converter.convert_to_excel(pdf_path, page_num, output_path)


def convert_pdf_page_to_text_for_llm(
    pdf_path: Path,
    page_num: int,
    method: str = "pdfplumber",
    temp_dir: Optional[Path] = None,
    keep_excel_file: bool = False
) -> str:
    """
    PDF 페이지를 실제 엑셀 파일로 변환한 후, 그 파일을 읽어서 텍스트로 변환 (간편 함수)
    
    프로세스:
    1. PDF → 엑셀 파일(.xlsx) 생성
    2. 엑셀 파일을 pandas로 읽기
    3. 읽은 내용을 텍스트로 변환
    4. 그 텍스트를 반환 (LLM에 전달)
    
    Args:
        pdf_path: PDF 파일 경로
        page_num: 페이지 번호 (1부터 시작)
        method: 변환 방법 ("tabula", "pdfplumber", "pymupdf")
        temp_dir: 엑셀 파일 저장 디렉토리
        keep_excel_file: 엑셀 파일을 유지할지 여부
        
    Returns:
        LLM에 전달할 수 있는 텍스트 형식 (엑셀 파일에서 읽은 내용)
    """
    converter = PdfToExcelConverter(method=method)
    return converter.convert_to_text_for_llm(
        pdf_path, 
        page_num, 
        temp_dir=temp_dir,
        keep_excel_file=keep_excel_file
    )

