"""
docs/현업테스트_기능별비교체크리스트.xlsx 생성.
실행: uv run python scripts/export_uat_checklist_xlsx.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "docs" / "현업테스트_기능별비교체크리스트.xlsx"

FILL_BLUE   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FILL_ORANGE = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
FILL_GRAY   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_GREEN  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_CAT = {
    "비교기준":   PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "인증":       PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
    "대시보드":   PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "업로드":     PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid"),
    "검토탭":     PatternFill(start_color="EAD1DC", end_color="EAD1DC", fill_type="solid"),
    "그리드":     PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid"),
    "정답지":     PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid"),
    "SAP엑셀":   PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
    "관리자":     PatternFill(start_color="D0E0E3", end_color="D0E0E3", fill_type="solid"),
    "동시접속":   PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid"),
    "예외상황":   PatternFill(start_color="E6B8AF", end_color="E6B8AF", fill_type="solid"),
}

FONT_WHITE_BOLD = Font(bold=True, color="FFFFFF")
FONT_BLACK_BOLD = Font(bold=True)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN = Side(style="thin")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 통합 컬럼 헤더
MAIN_HEADERS = ["분류", "확인 항목", "어떻게 확인하나", "구체적 포인트 / 기대 결과", "비고", "확인 (O/X)"]
COL_WIDTHS   = [14,     52,          38,                 42,                            22,     10]

# ── 각 분류별 rows: [확인 항목, 어떻게 확인하나, 구체적 포인트, 비고] ──────────────

ROWS: list[tuple[str, list[str]]] = []

def _add(category: str, items: list[list[str]]) -> None:
    for item in items:
        # item = [확인 항목, 어떻게 확인하나, 구체적 포인트, 비고(optional)]
        while len(item) < 4:
            item.append("")
        ROWS.append((category, item))


_add("비교기준", [
    ["원문 PDF — 추출 행 수·값 1차 기준",
     "업로드한 조건청구서 파일 직접 확인",
     "시스템 추출값과 직접 대조하는 기준",
     ""],
    ["지난달 수기 작업 엑셀 — 숫자·코드 정답 기준",
     "같은 PDF로 사람이 직접 만든 정리본",
     "그리드·SAP 엑셀 값과 일치 여부 비교",
     ""],
    ["SAP 업로드 양식 — 열 순서·코드 체계",
     "회사 표준 SAP 업로드 엑셀 형식",
     "규칙에 맞는 열 구성·계산식 여부",
     ""],
    ["마스터 파일 (소매처·판매처·상품·단가 CSV)",
     "database/csv/ 기준 데이터",
     "이름·코드 자동 매핑이 올바른지",
     ""],
])

_add("인증", [
    ["올바른 아이디·비밀번호로 로그인",
     "부여된 계정으로 로그인 시도",
     "로그인 성공·화면 전환",
     ""],
    ["틀린 비밀번호 입력 시 안내 문구 표시",
     "비밀번호 한 글자 틀리게 입력",
     "오류 메시지 표시 여부",
     ""],
    ["로그아웃 후 재접속",
     "로그아웃 → 다시 로그인",
     "세션 끊김·이전 작업 데이터 유지 여부",
     ""],
    ["비밀번호 변경 후 새 비밀번호로 로그인",
     "비밀번호 변경 → 로그아웃 → 재로그인",
     "변경 전 비밀번호 거부 여부",
     ""],
    ["일반 계정 접속 시 관리자 탭 보이지 않음",
     "일반 계정 로그인 후 탭 목록 확인",
     "관리자 탭이 없거나 접근 막혀야 함",
     ""],
    ["오랫동안 방치 후 자동 로그아웃 여부",
     "30분 이상 방치 후 화면 조작",
     "세션 만료 안내 문구 표시",
     ""],
])

_add("대시보드", [
    ["연월 필터가 올바른 달 데이터를 보여주는지",
     "이번 청구 월로 필터 변경",
     "건수·그래프가 직관과 일치하는지",
     ""],
    ["검토 진행률이 실제 체크 수와 맞는지",
     "검토 탭 실제 체크 수와 비교",
     "숫자 일치 여부",
     ""],
    ["거래처별 요약 집계가 맞는지",
     "특정 거래처 건수·금액을 수기 엑셀과 비교",
     "상위 1~2개 거래처 샘플로 확인",
     ""],
    ["이전 달 데이터가 이번 달로 섞이지 않는지",
     "연월 필터를 지난달로 바꿔 확인",
     "이번 달 문서가 지난달 필터에 보이면 오류",
     ""],
])

_add("업로드", [
    ["채널 선택 (FINET / MAIL) 후 업로드",
     "Excel변換 채널과 OCR 채널 각각 파일 업로드",
     "채널에 따라 처리 방식이 다름 (FINET=엑셀변환, MAIL=OCR). 채널 표시 확인",
     ""],
    ["청구 연월 선택 필수 — 연월 미선택 시 업로드 불가",
     "연월을 선택하지 않고 업로드 시도",
     "업로드 버튼이 비활성이거나 안내 문구 표시. 연·월 모두 선택해야 업로드 가능",
     ""],
    ["PDF 파일 1개 선택 후 업로드 완료",
     "파일 선택 → 업로드 → 목록 확인",
     "업로드 후 목록에 파일명·페이지 수 표시",
     ""],
    ["PDF 여러 개 동시 업로드",
     "3~5개 파일 한 번에 선택 후 업로드",
     "모두 목록에 정상 표시",
     ""],
    ["업로드 진행 상황 (페이지 단위 진행률) 표시",
     "업로드 중 화면 관찰",
     "'처리 중: 3/10' 같은 페이지 진행률 표시 여부. 10분 넘으면 타임아웃 안내",
     ""],
    ["PDF 아닌 파일 업로드 시도 시 오류 안내",
     "jpg, xlsx 파일 선택 후 업로드",
     "오류 안내 문구 표시. 그냥 통과되면 오류",
     ""],
    ["동일 파일명 재업로드 시 안내 여부",
     "이미 올린 파일과 같은 이름 파일 재업로드",
     "'이미 존재합니다' 등 안내 여부",
     ""],
    ["업로드 완료 목록에서 원하는 파일 찾기",
     "거래처명·연월로 필터·검색",
     "원하는 파일이 빠르게 찾아지는지",
     ""],
    ["PDF 미리보기(원본 이미지) 표시",
     "파일 클릭 → 미리보기 확인",
     "이미지가 선명하게, 올바른 페이지로 열리는지",
     ""],
    ["양식 유형(01~05) 자동 분류 결과 확인",
     "업로드 완료 후 문서 목록에서 양식 유형 표시 확인",
     "PDF 내용에 맞는 유형으로 자동 분류됐는지. 잘못 분류된 경우 비고에 기록",
     ""],
    ["잘못 올린 문서 삭제",
     "업로드 완료 목록에서 삭제 시도",
     "삭제 후 목록에서 사라지는지. 분석 데이터도 함께 삭제되는지",
     ""],
    ["페이지가 많은 문서 (20페이지 이상) 업로드·분석",
     "페이지 많은 PDF 업로드 후 분석 실행",
     "누락 페이지 없이 전부 분석 완료되는지. 시간이 과도하게 걸리지 않는지",
     ""],
])

_add("검토탭", [
    ["거래처 검색 시 해당 거래처 문서만 표시",
     "특정 거래처명 검색",
     "다른 거래처 문서가 섞이면 오류",
     ""],
    ["'담당 거래처만 보기' 필터",
     "담당取引先 선택 버튼 클릭 → 자기 거래처 선택",
     "선택한 거래처 문서만 표시. 다른 사람 거래처 데이터 안 보여야 함",
     ""],
    ["연월·양식 유형 필터 동작",
     "이번 달·특정 유형으로 필터",
     "해당 조건 문서만 표시",
     ""],
    ["1차/2차 검토 필터 — 세분화 확인",
     "1次완료/1次미완료/2次완료/2次미완료 각각 필터 적용",
     "각 필터가 정확하게 동작. 이미 체크한 건 사라지고, 미체크만 나오는지",
     ""],
    ["문서 클릭 시 해당 PDF 원본이 옆에 표시",
     "검토 행 클릭",
     "같은 문서·같은 페이지가 오른쪽에 열려야 함",
     ""],
    ["페이지 이동 동작",
     "여러 페이지 문서에서 페이지 번호 이동",
     "해당 페이지 내용이 오른쪽 이미지와 일치",
     ""],
    ["표지/상세/요약 페이지 구분 (페이지 역할)",
     "여러 페이지 문서에서 표지·요약 페이지 확인",
     "표지·요약 페이지는 그리드 데이터 행이 없어야 정상. 상세(detail) 페이지만 행이 표시",
     ""],
    ["분석 진행 중인 문서 '분석 중' 표시",
     "분석 시작 직후 검토 탭 확인",
     "'분석 중' 표시 없으면 오류 가능성",
     ""],
    ["이미 분석한 문서를 다시 분석(재분석)",
     "분석 완료 문서에서 재분석 실행",
     "이전 결과가 새 결과로 정상 덮어써지는지",
     ""],
    ["분석 소요 시간 기록",
     "분석 시작~완료까지 시간 측정 (체감)",
     "비고 란에 대략 몇 분 걸렸는지 기록해 주세요",
     ""],
    ["정답지 작성(解答作成) 진입",
     "해당 버튼 클릭",
     "정답지 탭으로 이동 여부",
     ""],
])

_add("그리드", [
    ["[추출] 행 수가 PDF 실제 행 수와 일치",
     "PDF 표 행 수를 직접 세어 그리드 행 수와 비교",
     "누락·중복 없어야 함. 페이지 경계 행 특히 확인",
     ""],
    ["[추출] 거래처명·거래처 코드",
     "PDF 표기 그대로인지",
     "0(숫자)과 O(알파벳), 1과 l 혼동 오류 多",
     ""],
    ["[추출] 상품명·상품 코드",
     "PDF 및 수기 엑셀과 비교",
     "품번 자리수·공백 확인",
     ""],
    ["[추출] 수량·단위 (個/CS/ケース 등)",
     "PDF 수량 칸 직접 확인",
     "단위가 個인지 CS인지에 따라 숫자 의미 다름",
     ""],
    ["[추출] 금액·세율·합계",
     "PDF 소계·합계 vs 그리드 금액 컬럼",
     "양식별 금액 컬럼명 다름 (金額·請求金額·請求合計額)",
     ""],
    ["[추출] 이중 조건(条件2·金額2) 처리",
     "조건이 2개인 PDF로 확인",
     "조건2·금액2 컬럼 값 정확 여부",
     ""],
    ["[편집] 셀 수정 후 저장 → 새로고침 후 유지",
     "임의 셀 값 수정 → 저장(Ctrl+S 또는 保存 버튼) → 새로고침",
     "새로고침 후에도 수정값 유지. Ctrl+S로도 저장되는지 확인",
     ""],
    ["[편집] 저장 안 하고 실수로 새로고침(F5) 했을 때",
     "셀 수정 후 저장 누르지 않고 F5 또는 브라우저 새로고침",
     "'저장하지 않은 내용이 있습니다' 경고가 뜨는지. 안 뜨면 데이터 유실 위험",
     ""],
    ["[편집] 행 추가 (연필 메뉴 → ➕ 추가)",
     "연필 버튼 클릭 → 추가 선택 → 빈 행이 아래에 생기는지",
     "추가된 행에 값 입력 후 저장·새로고침해도 유지되는지",
     ""],
    ["[편집] 행 삭제",
     "연필 메뉴 → 삭제 선택",
     "삭제 전 확인 팝업 있는지. 삭제 후 새로고침해도 사라져 있는지",
     ""],
    ["[편집] 1차/2차 체크 및 담당자 기록",
     "1차 체크 → 마우스를 체크 위에 올려 툴팁 확인",
     "체크한 사람 이름·시각이 표시되는지",
     ""],
    ["[편집] 1차/2차 일괄 체크 (페이지 단위)",
     "페이지 상단 전체 체크 버튼으로 한 번에 체크",
     "해당 페이지 모든 행이 일괄 체크. 일부만 체크된 상태도 표시",
     ""],
    ["[편집] 두 명이 같은 행 수정 시 잠금 표시",
     "A가 셀 편집 중일 때 B가 같은 행 클릭",
     "B 화면에 '편집 중: A이름' 표시. A가 편집 완료 후 B가 수정 가능",
     ""],
    ["[편집] 컬럼 순서 드래그 변경 → 재접속 후 유지",
     "컬럼 헤더를 드래그해서 순서 변경 → 로그아웃 후 재로그인",
     "변경한 컬럼 순서가 유지되는지",
     ""],
    ["[단가] 단가 자동 매핑 후보 표시",
     "연필 메뉴 → 単価 선택 → 후보 목록 확인",
     "unit_price 마스터에 있는 코드라면 후보 표시. 선택 후 값 반영",
     ""],
    ["[첨부] 행에 첨부 플래그 표시",
     "연필 메뉴 → 添付 선택 → 첨부 표시",
     "첨부 표시된 행은 왼쪽 테두리 색 등으로 구분 가능한지",
     ""],
    ["[표시] 가격·코드 컬럼 숨기기/보이기 토글",
     "仕切·本部長·NET 등 가격 컬럼 숨기기 체크박스 클릭",
     "체크 해제하면 해당 컬럼 숨겨지는지. 재접속 후에도 설정 유지",
     ""],
])

_add("정답지", [
    ["정답지 탭 접근 (解答作成 버튼)",
     "검토 탭에서 해당 버튼 클릭",
     "정답지 화면으로 이동 여부",
     ""],
    ["왼쪽 PDF 이미지가 올바른 페이지 표시",
     "정답지에서 페이지 이동",
     "PDF와 오른쪽 입력 항목이 같은 페이지를 가리키는지",
     ""],
    ["정답 값 직접 입력 후 저장 → 재진입 시 유지",
     "항목 입력 → 저장 → 탭 닫고 재진입",
     "저장 후 재진입 시 값 유지",
     ""],
    ["저장한 정답이 검토 그리드에도 반영",
     "정답지 저장 후 검토 탭 확인",
     "동일 문서·페이지 데이터 업데이트 여부",
     ""],
    ["학습 요청(RAG 반영) 동작",
     "학습 요청 버튼 클릭",
     "대시보드 RAG 수치 변화 여부",
     ""],
    ["두 명이 같은 정답지 동시 저장",
     "A·B 동시에 같은 정답지 수정 후 저장",
     "저장 순서에 따른 결과 확인",
     ""],
])

_add("SAP엑셀", [
    ["연월 선택 시 해당 월 대상 파일 목록·건수 표시",
     "SAP 탭에서 연월 변경",
     "양식별 파일명·행 수가 표시. 검토 탭 데이터와 건수 일치",
     ""],
    ["미리보기 토글 → 실제 엑셀 내용 사전 확인",
     "미리보기 버튼 클릭 → 양식별 필터로 내용 확인",
     "다운로드 전에 어떤 데이터가 들어가는지 미리 볼 수 있어야 함",
     ""],
    ["다운로드 버튼 클릭 시 엑셀 파일 생성·저장",
     "다운로드 버튼 클릭",
     "파일이 열리는지, 0바이트 아닌지",
     ""],
    ["W열 — 대상 연월 (예: 2025.01)",
     "화면 선택 연월 = 파일 W열",
     "날짜 형식 일치 여부",
     ""],
    ["B열 — 판매처명 (受注先코드 → sap_retail 매핑)",
     "C열 코드로 판매처명 찾아 B열과 비교",
     "코드 없으면 공란 또는 오류 메시지",
     ""],
    ["C열 — 受注先코드",
     "PDF 거래처 코드 = C열",
     "자리수·앞자리 0 유지",
     ""],
    ["J열 — 小売先코드 / K열 — 소매처명",
     "PDF 소매처 코드 = J열, K열 매핑 결과 확인",
     "J열 코드로 소매처명 찾아 K열 비교",
     ""],
    ["L열 — 상품명 / M열 — 商品코드",
     "M열 코드로 상품명 찾아 L열 비교",
     "코드 자리수·공백 확인",
     ""],
    ["N열·O열 — 2합환산·단일상자환산 단가",
     "M열 코드 → unit_price 매핑 결과",
     "단가가 수기 엑셀 단가와 일치하는지",
     ""],
    ["T열 — 수량 (양식별 규칙 다름)",
     "유형01: 단위=個→数量, CS→入数×数量\n유형02~05는 각 양식 규칙 확인",
     "유형별로 각각 확인",
     "⚠️ 개발 진행 중"],
    ["AL열 — 금액 (양식별 규칙 다름)",
     "유형01·02: 金額 / 유형03: 請求金額\n유형04: 金額+金額2 / 유형05: 請求合計額",
     "이중 조건 있으면 합산 여부 확인",
     "⚠️ 개발 진행 중"],
    ["U열 — 엑셀 수식 삽입 (=P×N + R×O + T)",
     "엑셀 열어서 U열 셀 클릭 시 수식 표시",
     "값이 아닌 수식으로 들어가야 함",
     "⚠️ 개발 진행 중"],
    ["엑셀 미리보기 행 수 = 실제 다운로드 행 수",
     "미리보기 행 수와 파일 행 수 비교",
     "",
     ""],
    ["다운로드 파일명이 구분 가능한지",
     "여러 건 다운로드 후 파일명 확인",
     "연월·거래처 등으로 구분 가능. 파일이 여러 개일 때 어느 건인지 알 수 있어야 함",
     ""],
    ["검토 탭에서 수정한 값이 SAP 엑셀에도 반영",
     "그리드에서 셀 수정·저장 → SAP 엑셀 다운로드 → 해당 셀 확인",
     "수정한 값이 그대로 엑셀에 들어가야 함",
     ""],
])

_add("관리자", [
    ["마스터 파일 조회·검색·다운로드",
     "각 마스터 탭(소매처·판매처·상품·단가) 클릭 → 검색 → 엑셀 다운로드",
     "데이터 표시·검색·다운로드 모두 동작하는지",
     "관리자 전용"],
    ["마스터 CSV/엑셀 파일 업로드 (교체)",
     "새 CSV 또는 XLSX 파일 업로드",
     "기존 데이터가 새 파일로 교체. 매핑(상품명·단가 등)에 반영되는지",
     "관리자 전용"],
    ["사용자 계정 생성·수정·삭제",
     "새 사용자 추가 → 정보 수정 → 삭제",
     "생성한 계정으로 로그인 가능. 삭제한 계정 로그인 불가",
     "관리자 전용"],
    ["비밀번호 초기화",
     "특정 사용자 비밀번호 초기화(初期化) 버튼 클릭",
     "해당 사용자가 초기 비밀번호로 로그인 가능",
     "관리자 전용"],
    ["관리자 권한 부여·해제",
     "일반 사용자에게 관리자 체크 → 저장",
     "해당 사용자 재로그인 시 관리자 탭 보이는지",
     "관리자 전용"],
    ["RAG 학습 현황 반영",
     "정답지 저장 후 RAG 수치 확인",
     "대시보드 RAG 통계와 연계",
     "관리자 전용"],
])

_add("예외상황", [
    ["분석 중 브라우저를 닫거나 다른 페이지로 이동",
     "분석 시작 후 브라우저 닫기 → 다시 열기",
     "돌아왔을 때 분석이 계속 진행 중인지, 중단됐는지",
     ""],
    ["인터넷이 잠깐 끊겼다가 복구됐을 때",
     "Wi-Fi 끄기 → 5초 후 켜기 → 화면 조작",
     "에러 메시지가 뜨는지, 자동으로 다시 연결되는지",
     ""],
    ["사용 중인 브라우저(Chrome / Edge 등)에서 화면 깨짐 없음",
     "현업에서 실제 사용하는 브라우저로 전체 기능 확인",
     "레이아웃·글자 깨짐·버튼 안 눌림 등 없어야 함",
     "비고에 브라우저명 기록"],
])

_add("동시접속", [
    ["7명 이상 동시 로그인",
     "7~9명이 동시에 로그인",
     "각자 정상 접속. 한 명 접속이 다른 사람에게 영향 없어야 함",
     "전체 인원"],
    ["각자 다른 문서 작업 시 데이터 섞임 없음",
     "A가 보는 문서와 B가 보는 문서가 서로 화면에 안 보임",
     "A 그리드에 B 문서 데이터가 나오면 오류",
     "2명 이상"],
    ["여러 명이 동시에 분석 실행",
     "3명이 각자 다른 파일 분석 동시 시작",
     "각자 분석 정상 완료. 서버 오류·지연 없는지",
     "3명"],
    ["같은 행을 두 명이 동시 수정",
     "A·B 동시에 같은 셀 다른 값 입력 후 저장",
     "마지막 저장 반영 또는 충돌 안내. 값 사라지면 오류",
     "2명"],
    ["같은 정답지를 두 명이 동시 저장",
     "A·B 동시에 같은 정답지 수정 후 저장",
     "저장 순서에 따른 결과 확인",
     "2명"],
    ["여러 명이 동시에 SAP 엑셀 다운로드",
     "3명이 동시에 같은 연월 SAP 엑셀 다운로드",
     "각자 파일 정상 수령. 파일 내용 일치",
     "3명"],
    ["분석·편집·다운로드 동시 진행",
     "A:분석, B:그리드 편집, C:SAP 다운로드 동시",
     "각 작업 정상 완료. 화면 느려지거나 오류 없는지",
     "3명"],
])


# ── 시트 생성 함수들 ──────────────────────────────────────────────────────────

def _autosize(ws, col_widths: list[int]) -> None:
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_main(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "체크리스트"

    # 헤더
    for c, (h, w) in enumerate(zip(MAIN_HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = FONT_WHITE_BOLD
        cell.fill = FILL_BLUE
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    current_cat = None
    for r, (cat, item) in enumerate(ROWS, 2):
        # 분류 열
        cat_cell = ws.cell(row=r, column=1, value=cat if cat != current_cat else "")
        cat_cell.alignment = CENTER
        cat_cell.border = THIN_BORDER
        if cat in FILL_CAT:
            cat_cell.fill = FILL_CAT[cat]
        if cat != current_cat:
            cat_cell.font = FONT_BLACK_BOLD
            current_cat = cat

        # 확인 항목·어떻게·포인트·비고
        for c, val in enumerate(item, 2):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = WRAP_TOP
            cell.border = THIN_BORDER
            if cat in FILL_CAT:
                cell.fill = FILL_CAT[cat]
            # SAP 개발진행중 행 노란색 강조
            if val.startswith("⚠️"):
                for cc in range(1, len(MAIN_HEADERS) + 1):
                    ws.cell(row=r, column=cc).fill = FILL_YELLOW

        # 확인 열 (마지막)
        chk = ws.cell(row=r, column=len(MAIN_HEADERS), value="")
        chk.alignment = CENTER
        chk.border = THIN_BORDER

    _autosize(ws, COL_WIDTHS)

    # SAP 미완성 안내
    note_r = len(ROWS) + 3
    note = ws.cell(row=note_r, column=1,
                   value="⚠️ 노란색 행(SAP엑셀 일부) = 현재 개발 진행 중. 이번 테스트에서 결과가 다를 수 있으며 검증 생략 가능.")
    note.font = Font(bold=True, color="CC0000")
    note.alignment = WRAP_TOP
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=len(MAIN_HEADERS))


def sheet_ai_error(wb: Workbook) -> None:
    ws = wb.create_sheet("AI추출_오류기록")

    guide_lines = [
        "AI가 잘못 추출한 경우 이 시트에 기록해 주세요.",
        "",
        '좋은 예시: "A문서 3페이지, 金額 열 → 수기 정답: 1,692  /  시스템 추출: 1,629  (숫자 순서 바뀜)"',
        '나쁜 예시: "학습이 안 돼요" → 개발팀이 대응 불가능',
        "",
        "아래 표에 한 건씩 작성해 주세요. 가능하면 어떤 PDF의 몇 페이지인지, 어느 열인지 적어주세요.",
    ]
    for r, line in enumerate(guide_lines, 1):
        cell = ws.cell(row=r, column=1, value=line)
        if r == 1:
            cell.font = Font(bold=True, size=12)
        elif r in (3, 4):
            cell.font = Font(italic=True, color="595959")
        cell.alignment = WRAP_TOP
    ws.column_dimensions["A"].width = 100
    ws.row_dimensions[3].height = 22

    header_row = len(guide_lines) + 2
    headers = [
        "문서명 (파일명)",
        "페이지\n번호",
        "양식\n유형\n(1~5)",
        "항목명\n(컬럼)",
        "수기 정답값\n(기대값)",
        "시스템\n추출값",
        "일치\n(O/X)",
        "오류 패턴\n(예: 숫자 순서 바뀜, 행 누락 등)",
        "비고",
    ]
    col_widths = [30, 8, 8, 20, 18, 18, 8, 40, 25]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = FONT_WHITE_BOLD
        cell.fill = FILL_ORANGE
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row].height = 42

    samples = [
        ["(예시) 문서A.pdf", "3", "1", "金額", "1,692", "1,629", "X",
         "숫자 순서 바뀜 (OCR 오인식)", ""],
        ["(예시) 문서B.pdf", "1", "2", "JAN코드", "8801043019217", "8801043091217", "X",
         "가운데 숫자 위치 바뀜", "스캔 품질 불량 가능성"],
    ]
    for r, row in enumerate(samples, header_row + 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = WRAP_TOP
            cell.border = THIN_BORDER
            cell.fill = FILL_GREEN

    for r in range(header_row + len(samples) + 1, header_row + 20):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c, value="")
            cell.border = THIN_BORDER
            cell.alignment = WRAP_TOP

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_ux_feedback(wb: Workbook) -> None:
    ws = wb.create_sheet("화면_불편사항")

    guide = (
        "화면 사용 중 불편하거나 이상한 점을 자유롭게 기록해 주세요.\n"
        "예) 버튼을 찾기 어렵다 / 오류가 생겼는데 아무 메시지가 없다 / 글씨가 너무 작다 / "
        "저장했는데 어디에 저장됐는지 모르겠다 등"
    )
    cell = ws.cell(row=1, column=1, value=guide)
    cell.font = Font(size=11)
    cell.alignment = WRAP_TOP
    ws.row_dimensions[1].height = 50

    headers = ["화면·기능 위치", "불편하거나 이상한 점", "심각도\n(높음/보통/낮음)", "담당자", "비고"]
    col_widths = [25, 50, 15, 12, 20]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = FONT_WHITE_BOLD
        cell.fill = FILL_GRAY
        cell.font = Font(bold=True)
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for r in range(4, 22):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c, value="")
            cell.border = THIN_BORDER
            cell.alignment = WRAP_TOP

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_scenario(wb: Workbook) -> None:
    ws = wb.create_sheet("통합_시나리오")
    headers = ["시나리오", "단계별 흐름", "확인 (O/X)"]
    rows = [
        ["시나리오 1\n[표준 1건 전체 흐름]",
         "익숙한 거래처 PDF 1개\n→ 업로드\n→ 분석 완료 대기\n→ 검토 탭에서 수기 엑셀과 전항목 대조\n→ 수정 필요 시 셀 편집·저장\n→ SAP 엑셀 다운로드\n→ 수기 엑셀 핵심 열(금액·코드·수량)과 비교"],
        ["시나리오 2\n[난이도 높은 문서]",
         "스캔 품질 낮거나 페이지 많은 PDF\n→ 누락·오인식·페이지 경계 행 확인\n→ 이중 조건(조건2·금액2) 있는 문서 포함 권장"],
        ["시나리오 3\n[2인 검수 프로세스]",
         "A가 1차 체크 완료\n→ B가 2차 체크\n→ SAP 엑셀에 해당 행 반영 확인"],
        ["시나리오 4\n[동시 접속 7명]",
         "7명 동시 로그인\n→ 각자 다른 문서 분석 동시 시작\n→ 분석 완료 후 각자 그리드 검토\n→ 3명이 동시에 SAP 엑셀 다운로드"],
        ["시나리오 5\n[정답지 저장 → 재분석]",
         "정답지 탭에서 특정 문서 정답 입력·저장\n→ 학습 요청\n→ 같은 문서 재분석\n→ 이전보다 추출 개선 여부 확인"],
        ["시나리오 6\n[마스터 교체 → 매핑 변화]",
         "(관리자) 마스터 CSV 교체 업로드\n→ SAP 엑셀 다운로드\n→ B~O열 이름·단가가 새 마스터 반영 확인"],
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = FONT_WHITE_BOLD
        cell.fill = FILL_BLUE
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 22

    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = WRAP_TOP
            cell.border = THIN_BORDER
        ws.cell(row=r, column=3, value="").border = THIN_BORDER

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 65
    ws.column_dimensions["C"].width = 12


def main() -> None:
    wb = Workbook()

    sheet_main(wb)
    sheet_ai_error(wb)
    sheet_ux_feedback(wb)
    sheet_scenario(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"완료: {OUT}")


if __name__ == "__main__":
    main()
