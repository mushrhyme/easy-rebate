"""
RAG / 벡터 DB 관리용 관리자 API
"""
import asyncio
import csv
import io
import logging
import re
from pathlib import Path
from typing import Optional, Dict, Any, List, Tuple
import json

logger = logging.getLogger(__name__)

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from psycopg2.extras import RealDictCursor

from database.registry import get_db
from backend.core.auth import get_current_user
from backend.api.routes.documents import request_cancel_reanalysis_for_document, _ensure_phase1_rag_columns
from modules.core.build_faiss_db import build_faiss_db
from modules.core.rag_manager import get_rag_manager
from modules.utils.hash_utils import compute_page_hash, get_page_key
from modules.utils.config import get_project_root


router = APIRouter()


def _ensure_rag_candidate_column(db):
    """
    is_rag_candidate 컬럼이 없으면 추가 (마이그레이션)
    """
    try:
        with db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT column_name FROM information_schema.columns
                WHERE table_name = 'page_data_current' AND column_name = 'is_rag_candidate'
            """)
            if cursor.fetchone():
                return
            cursor.execute("ALTER TABLE page_data_current ADD COLUMN IF NOT EXISTS is_rag_candidate BOOLEAN NOT NULL DEFAULT FALSE")
            cursor.execute("ALTER TABLE page_data_archive ADD COLUMN IF NOT EXISTS is_rag_candidate BOOLEAN NOT NULL DEFAULT FALSE")
            conn.commit()
    except Exception:
        pass


def _ensure_last_edited_columns(db):
    """last_edited_at, last_edited_by_user_id 컬럼이 없으면 추가 (Phase 1 마이그레이션)"""
    try:
        with db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT column_name FROM information_schema.columns
                WHERE table_name = 'page_data_current' AND column_name = 'last_edited_at'
            """)
            if cursor.fetchone():
                return
            cursor.execute("ALTER TABLE page_data_current ADD COLUMN IF NOT EXISTS last_edited_at TIMESTAMPTZ NULL")
            cursor.execute("ALTER TABLE page_data_current ADD COLUMN IF NOT EXISTS last_edited_by_user_id INTEGER NULL")
            cursor.execute("ALTER TABLE page_data_archive ADD COLUMN IF NOT EXISTS last_edited_at TIMESTAMPTZ NULL")
            cursor.execute("ALTER TABLE page_data_archive ADD COLUMN IF NOT EXISTS last_edited_by_user_id INTEGER NULL")
            conn.commit()
    except Exception:
        pass


def _ensure_admin(user: Dict[str, Any]) -> None:
    """
    관리자 권한 확인 (username='admin' 또는 is_admin=True)
    """
    if user.get("username") != "admin" and not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="관리자만 접근할 수 있습니다")


def _extract_ocr_for_page(db, pdf_filename: str, page_number: int) -> str:
    """
    학습 리퀘스트 시 임베딩용 OCR 추출.
    우선순위: page_meta._ocr_text → page_data_current.ocr_text (DB만 사용, debug2 미사용).
    """
    pdf_name = pdf_filename[:-4] if pdf_filename.lower().endswith(".pdf") else pdf_filename

    # 0) DB: page_meta._ocr_text 또는 ocr_text 컬럼
    try:
        with db.get_connection() as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT page_meta, ocr_text FROM page_data_current WHERE pdf_filename = %s AND page_number = %s",
                (pdf_filename, page_number),
            )
            row = cur.fetchone()
            if row:
                meta = row[0] if isinstance(row[0], dict) else (json.loads(row[0]) if row[0] and isinstance(row[0], str) else None)
                if isinstance(meta, dict) and (meta.get("_ocr_text") or "").strip():
                    return (meta["_ocr_text"] or "").strip()
                if row[1] and (row[1] or "").strip():
                    return (row[1] or "").strip()
    except Exception:
        pass

    return ""


@router.post("/build")
async def build_vector_db(
    payload: Dict[str, Optional[str]] = None,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    벡터 DB 재구축 트리거 (관리자 전용)

    - 선택적으로 form_type(예: '01', '02') 를 지정할 수 있음
      지정하지 않으면 모든 양식 폴더 대상
    - 현재는 img 폴더 기반 build_faiss_db 를 그대로 사용
    """
    _ensure_admin(current_user)

    form_type: Optional[str] = None
    if payload:
        form_type = payload.get("form_type") or None

    # 스레드 풀에서 실행하여 이벤트 루프 블로킹 방지
    logger.info("ベクターDB再構築 開始 (img フォルダ走査)")
    try:
        await asyncio.to_thread(
            build_faiss_db,
            form_folder=form_type,
            auto_merge=True,
            text_extraction_method="excel",
        )
        logger.info("ベクターDB再構築 完了")
    except Exception as e:
        logger.exception("ベクターDB再構築 エラー")
        raise HTTPException(status_code=500, detail=f"벡터 DB 생성 중 오류가 발생했습니다: {e}")

    # 최신 통계 조회
    rag_manager = get_rag_manager()
    total_vectors = rag_manager.count_examples()

    def _fetch_per_form(db):
        with db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT form_type, SUM(vector_count) AS total_vectors FROM rag_vector_index GROUP BY form_type ORDER BY form_type")
            return [{"form_type": row[0], "vector_count": int(row[1] or 0)} for row in cursor.fetchall()]
    db = get_db()
    per_form = await db.run_sync(_fetch_per_form, db)
    return {"success": True, "message": "벡터 DB 생성이 완료되었습니다.", "total_vectors": int(total_vectors), "per_form_type": per_form}


@router.get("/status")
async def get_vector_db_status(
    year: Optional[int] = Query(None, description="선택 연월(년). 기간별 정답지 수에 사용"),
    month: Optional[int] = Query(None, description="선택 연월(월). 기간별 정답지 수에 사용"),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    현재 벡터 DB 상태 조회. 로그인 사용자 전체 허용 (읽기 전용).
    - total_vectors / merged_pages: rag_page_embeddings(pgvector) 기준. 테이블 없으면 0.
    - unused_pages: total_vectors - merged_pages
    - answer_key_pages_in_period: year/month 지정 시 해당 연월 merged 수 (pgvector 시 updated_at 기준)
    """
    rag_manager = get_rag_manager()

    def _fetch_status_sync(database, y, m):
        with database.get_connection() as conn:
            cursor = conn.cursor()
            # pgvector(rag_page_embeddings)가 있으면 여기 기준으로 집계 — 학습リクエスト 반영
            cursor.execute("""
                SELECT EXISTS (SELECT 1 FROM information_schema.tables WHERE table_schema = 'public' AND table_name = 'rag_page_embeddings')
            """)
            if not cursor.fetchone()[0]:
                return 0, [], 0, None
            cursor.execute("SELECT COUNT(*) FROM rag_page_embeddings")
            total_vectors = int(cursor.fetchone()[0] or 0)
            cursor.execute("""
                SELECT COALESCE(form_type, '') AS form_type, COUNT(*) AS cnt
                FROM rag_page_embeddings GROUP BY form_type ORDER BY form_type
            """)
            per_form = [{"form_type": (r[0] or "").strip() or None, "vector_count": int(r[1] or 0)} for r in cursor.fetchall()]
            merged_pages = total_vectors
            answer_key_pages_in_period = None
            if y is not None and m is not None:
                cursor.execute("""
                    SELECT COUNT(*) FROM rag_page_embeddings
                    WHERE updated_at >= date_trunc('month', make_date(%s, %s, 1))
                      AND updated_at < date_trunc('month', make_date(%s, %s, 1)) + interval '1 month'
                """, (y, m, y, m))
                answer_key_pages_in_period = int(cursor.fetchone()[0] or 0)
            return total_vectors, per_form, merged_pages, answer_key_pages_in_period

    db = get_db()
    total_vectors, per_form, merged_pages, answer_key_pages_in_period = await db.run_sync(_fetch_status_sync, db, year, month)
    unused_pages = max(0, total_vectors - merged_pages)
    return {
        "total_vectors": int(total_vectors),
        "per_form_type": per_form,
        "merged_pages": merged_pages,
        "unused_pages": unused_pages,
        "answer_key_pages_in_period": answer_key_pages_in_period,
    }


class BuildFromLearningPagesRequest(BaseModel):
    """검토 화면에서 선택한 페이지들로 벡터 DB 생성 요청"""

    form_type: Optional[str] = None


@router.get("/learning-flag")
async def get_learning_flag(
    pdf_filename: str,
    page_number: int,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    특정 페이지가 벡터 DB 학습 대상인지 여부 조회 (관리자 전용)
    """
    _ensure_admin(current_user)

    def _get_flag_sync(database, pdf: str, page: int):
        _ensure_rag_candidate_column(database)
        with database.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT is_rag_candidate FROM page_data_current WHERE pdf_filename = %s AND page_number = %s", (pdf, page))
            row = cursor.fetchone()
        return bool(row[0]) if row else False
    db = get_db()
    selected = await db.run_sync(_get_flag_sync, db, pdf_filename, page_number)
    return {"selected": selected}


class LearningFlagRequest(BaseModel):
    pdf_filename: str
    page_number: int
    selected: bool


@router.post("/learning-flag")
async def set_learning_flag(
    request: LearningFlagRequest,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    특정 페이지의 벡터 DB 학습 대상 플래그 설정 (관리자 전용)
    """
    _ensure_admin(current_user)

    def _set_flag_sync(database):
        _ensure_rag_candidate_column(database)
        with database.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                "UPDATE page_data_current SET is_rag_candidate = %s, updated_at = CURRENT_TIMESTAMP WHERE pdf_filename = %s AND page_number = %s",
                (request.selected, request.pdf_filename, request.page_number),
            )
            if cursor.rowcount == 0:
                raise HTTPException(status_code=404, detail="Page not found")
    db = get_db()
    await db.run_sync(_set_flag_sync, db)
    return {"success": True}


class LearningRequestPageRequest(BaseModel):
    """단일 페이지 학습 요청 (Phase 1)"""
    pdf_filename: str
    page_number: int


def _is_admin(user: Dict[str, Any]) -> bool:
    """관리자 여부 (username='admin' 또는 is_admin=True)."""
    return user.get("username") == "admin" or bool(user.get("is_admin"))


async def execute_learning_request_page(
    pdf_filename: str,
    page_number: int,
    current_user: Optional[Dict[str, Any]] = None,
) -> dict:
    """
    단일 페이지 학습 요청 실행. 로그인 사용자 전원 호출 가능 (관리자 체크 없음).
    검토·정답지 탭의 「学習リクエスト」에서 사용. /api/search/learning-request-page 에서도 호출.
    current_user가 관리자이면 성공 후 판매처·소매처 / 제품 RAG 정답지 벡터 인덱스 2개 자동 재구축.
    """
    db = get_db()
    _ensure_rag_candidate_column(db)
    _ensure_last_edited_columns(db)
    _ensure_phase1_rag_columns(db)
    request_cancel_reanalysis_for_document(pdf_filename)

    def _set_and_fetch_sync(database):
        with database.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                "UPDATE page_data_current SET is_rag_candidate = TRUE, updated_at = CURRENT_TIMESTAMP WHERE pdf_filename = %s AND page_number = %s",
                (pdf_filename, page_number),
            )
            if cursor.rowcount == 0:
                raise HTTPException(status_code=404, detail="Page not found")
        return _fetch_one_learning_page(database, pdf_filename, page_number)

    shard_page = await db.run_sync(_set_and_fetch_sync, db)
    if not shard_page:
        raise HTTPException(status_code=400, detail="벡터화할 수 있는 데이터가 없습니다(OCR 텍스트 등).")

    form_type = (shard_page.get("metadata") or {}).get("form_type") or None
    rag_manager = get_rag_manager()
    ok = await asyncio.to_thread(
        rag_manager.upsert_page_embedding,
        pdf_filename,
        page_number,
        shard_page.get("ocr_text", ""),
        shard_page.get("answer_json", {}),
        form_type,
    )
    if not ok:
        raise HTTPException(status_code=500, detail="pgvector 반영에 실패했습니다.")
    logger.info("학습 리퀘스트 반영: %s p.%s → pgvector(재분석 시 최신 벡터로 사용됨)", pdf_filename, page_number)
    try:
        with db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                "UPDATE documents_current SET current_vector_version = COALESCE(current_vector_version, 1) + 1 WHERE pdf_filename = %s",
                (pdf_filename,),
            )
            conn.commit()
    except Exception:
        pass

    # 관리자면 판매처·소매처 / 제품 RAG 정답지 벡터 인덱스 2개 재구축 (기준관리 탭 버튼과 동일)
    if current_user and _is_admin(current_user):
        try:
            rag = get_rag_manager()
            n_retail = rag.build_retail_rag_answer_index()
            n_product = rag.build_product_rag_answer_index()
            logger.info("학습 요청 후 RAG 정답지 인덱스 재구축: retail=%s, product=%s", n_retail, n_product)
        except Exception as e:
            logger.exception("학습 요청 후 RAG 정답지 인덱스 재구축 실패: %s", e)

    return {"success": True, "message": "해당 페이지를 벡터 DB에 반영했습니다."}


@router.post("/learning-request-page")
async def learning_request_page(
    body: LearningRequestPageRequest,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    해당 페이지만 벡터 DB에 반영. ※ 관리자 전용 아님. 로그인한 모든 사용자 호출 가능.
    (비관리자 403 방지를 위해 프론트는 /api/search/learning-request-page 사용 권장.)
    관리자일 경우 성공 후 판매처·소매처 / 제품 RAG 정답지 인덱스 자동 재구축.
    """
    return await execute_learning_request_page(body.pdf_filename, body.page_number, current_user)


def _fetch_one_learning_page(database, pdf_filename: str, page_number: int) -> Optional[Dict[str, Any]]:
    """단일 페이지에 대해 build-from-learning-pages와 동일한 shard_page 딕셔너리 생성.
    page_role이 detail인데 タイプ가 null인 행은 DB에 条件으로 저장한 뒤 벡터에 반영.
    """
    with database.get_connection() as conn:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("""
            SELECT p.pdf_filename, p.page_number, p.page_role, p.page_meta, i.item_id, i.item_order, i.item_data, d.form_type, d.data_year, d.data_month
            FROM page_data_current p
            LEFT JOIN items_current i ON i.pdf_filename = p.pdf_filename AND i.page_number = p.page_number
            LEFT JOIN documents_current d ON p.pdf_filename = d.pdf_filename
            WHERE p.pdf_filename = %s AND p.page_number = %s
            ORDER BY i.item_order NULLS LAST
        """, (pdf_filename, page_number))
        rows = cursor.fetchall()
        if not rows:
            return None
        first_row = rows[0]
        page_role = (first_row.get("page_role") or "detail").strip() or "detail"
        merged_items = []
        for row in rows:
            if row.get("item_id") is None:
                continue
            item_data = row.get("item_data") or {}
            if isinstance(item_data, str):
                try:
                    item_data = json.loads(item_data)
                except json.JSONDecodeError:
                    item_data = {}
            if not isinstance(item_data, dict):
                item_data = {}
            else:
                item_data = dict(item_data)
            # detail 페이지이고 タイプ가 null/빈값이면 DB에 条件으로 저장 (눈속임 아닌 실제 UPDATE)
            if page_role == "detail":
                _t = item_data.get("タイプ")
                if _t is None or (isinstance(_t, str) and not (_t or "").strip()):
                    item_data["タイプ"] = "条件"
                    cursor.execute(
                        "UPDATE items_current SET item_data = %s::json WHERE item_id = %s",
                        (json.dumps(item_data, ensure_ascii=False), row["item_id"]),
                    )
            merged_items.append(item_data)
        conn.commit()
    pdf_name = pdf_filename[:-4] if pdf_filename.lower().endswith(".pdf") else pdf_filename
    page_meta = first_row.get("page_meta") or {}
    if isinstance(page_meta, str):
        try:
            page_meta = json.loads(page_meta)
        except json.JSONDecodeError:
            page_meta = {}
    page_role = first_row.get("page_role") or "detail"
    ocr_text = _extract_ocr_for_page_sync(database, pdf_filename, page_number)
    if not ocr_text:
        return None
    answer_json = {**page_meta, "page_role": page_role, "items": merged_items}
    # タイプ 확정 전 저장된 page_meta 대비, 문서 확정 form_type으로 덮어써 벡터 DB에 항상 최신 반영
    answer_json["form_type"] = (first_row.get("form_type") or "").strip()
    metadata = {
        "pdf_name": pdf_name,
        "page_num": page_number,
        "form_type": (first_row.get("form_type") or "").strip(),
        "source": "db_learning_pages",
        "data_year": first_row.get("data_year"),
        "data_month": first_row.get("data_month"),
    }
    return {
        "pdf_name": pdf_name,
        "page_num": page_number,
        "ocr_text": ocr_text,
        "answer_json": answer_json,
        "metadata": metadata,
        "page_key": get_page_key(pdf_name, page_number),
        "page_hash": compute_page_hash(ocr_text, answer_json),
    }


def _extract_ocr_for_page_sync(database, pdf_filename: str, page_number: int) -> str:
    """동기 버전: OCR 추출 (page_meta._ocr_text → page_data_current.ocr_text, DB만 사용)."""
    try:
        with database.get_connection() as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT page_meta, ocr_text FROM page_data_current WHERE pdf_filename = %s AND page_number = %s",
                (pdf_filename, page_number),
            )
            row = cur.fetchone()
            if row:
                meta = row[0] if isinstance(row[0], dict) else (json.loads(row[0]) if row[0] and isinstance(row[0], str) else None)
                if isinstance(meta, dict) and (meta.get("_ocr_text") or "").strip():
                    return (meta["_ocr_text"] or "").strip()
                if row[1] and (row[1] or "").strip():
                    return (row[1] or "").strip()
    except Exception:
        pass
    return ""


@router.get("/learning-pages")
async def get_learning_pages(
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    현재 벡터 DB 학습 대상으로 체크된 페이지 목록 조회 (관리자 전용)
    """
    _ensure_admin(current_user)

    def _get_learning_pages_sync(database):
        _ensure_rag_candidate_column(database)
        with database.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT pdf_filename, page_number FROM page_data_current WHERE is_rag_candidate = TRUE ORDER BY pdf_filename, page_number")
            return [{"pdf_filename": r[0], "page_number": r[1]} for r in cursor.fetchall()]
    db = get_db()
    pages = await db.run_sync(_get_learning_pages_sync, db)
    return {"count": len(pages), "pages": pages}


@router.post("/build-from-learning-pages")
async def build_vector_from_learning_pages(
    request: BuildFromLearningPagesRequest,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    is_rag_candidate = TRUE 로 설정된 페이지들로부터 벡터 DB 생성 (관리자 전용)
    """
    _ensure_admin(current_user)

    def _fetch_learning_rows_sync(database):
        _ensure_rag_candidate_column(database)
        with database.get_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            cursor.execute("""
                SELECT p.pdf_filename, p.page_number, p.page_role, p.page_meta, i.item_id, i.item_order, i.item_data, d.form_type, d.data_year, d.data_month, d.document_metadata
                FROM page_data_current p
                LEFT JOIN items_current i ON i.pdf_filename = p.pdf_filename AND i.page_number = p.page_number
                LEFT JOIN documents_current d ON p.pdf_filename = d.pdf_filename
                WHERE p.is_rag_candidate = TRUE
                ORDER BY p.pdf_filename, p.page_number, i.item_order NULLS LAST
            """)
            return cursor.fetchall()
    db = get_db()
    rows: List[Dict[str, Any]] = await db.run_sync(_fetch_learning_rows_sync, db)
    if not rows:
        raise HTTPException(status_code=400, detail="학습 대상으로 선택된 페이지가 없습니다.")

    # (pdf_filename, page_number) 단위로 그룹화 (items 없는 페이지는 1행, item_id=NULL)
    from collections import defaultdict

    pages_map: Dict[Tuple[str, int], List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        key = (row["pdf_filename"], row["page_number"])
        pages_map[key].append(row)

    shard_pages: List[Dict[str, Any]] = []

    for (pdf_filename, page_number), item_rows in pages_map.items():
        # pdf_name (확장자 제거)
        pdf_name = pdf_filename[:-4] if pdf_filename.lower().endswith(".pdf") else pdf_filename

        merged_items: List[Dict[str, Any]] = []

        # 연월/폼타입 정보 (이 페이지가 속한 문서 기준: documents_current.form_type)
        first_row = item_rows[0]
        data_year = first_row.get("data_year")
        data_month = first_row.get("data_month")
        form_type_for_page = (first_row.get("form_type") or "").strip() or ""
        page_role = first_row.get("page_role") or "detail"
        page_meta_raw = first_row.get("page_meta")
        page_meta: Dict[str, Any] = {}
        if page_meta_raw is not None:
            if isinstance(page_meta_raw, dict):
                page_meta = dict(page_meta_raw)
            elif isinstance(page_meta_raw, str):
                try:
                    page_meta = json.loads(page_meta_raw)
                except json.JSONDecodeError:
                    pass

            for row in item_rows:
                if row.get("item_id") is None:
                    # items 없는 페이지(cover): item 행 스킵
                    continue

                item_data = row.get("item_data") or {}
                merged_item = dict(item_data) if isinstance(item_data, dict) else {}
                merged_items.append(merged_item)

        # 저장된 키 순서로 재정렬 (document_metadata.item_data_keys 사용)
        doc_meta = first_row.get("document_metadata") or {}
        if isinstance(doc_meta, dict):
            item_data_keys = doc_meta.get("item_data_keys")
            if item_data_keys and isinstance(item_data_keys, list) and merged_items:
                ordered_items = []
                for merged_item in merged_items:
                    ordered = {k: merged_item[k] for k in item_data_keys if k in merged_item}
                    ordered.update({k: merged_item[k] for k in merged_item if k not in item_data_keys})
                    ordered_items.append(ordered)
                merged_items = ordered_items

        # 실제 OCR 텍스트 추출 (페이지 이미지에서, 임베딩용)
        ocr_text = await asyncio.to_thread(_extract_ocr_for_page, db, pdf_filename, page_number)
        if not ocr_text:
            # OCR 추출 실패 시 스킵 (빈 텍스트로 임베딩할 수 없음)
            continue

        # answer_json: page_meta(문서 메타) + page_role + items. form_type은 문서 확정값으로 덮어쓰기
        answer_json: Dict[str, Any] = {
            **page_meta,
            "page_role": page_role,
            "items": merged_items,
            "form_type": form_type_for_page,
        }

        metadata: Dict[str, Any] = {
            "pdf_name": pdf_name,
            "page_num": page_number,
            "form_type": form_type_for_page,
            "source": "db_learning_pages",
            "data_year": data_year,
            "data_month": data_month,
        }

        page_key = get_page_key(pdf_name, page_number)
        page_hash = compute_page_hash(ocr_text, answer_json)

        shard_pages.append(
            {
                "pdf_name": pdf_name,
                "page_num": page_number,
                "ocr_text": ocr_text,
                "answer_json": answer_json,
                "metadata": metadata,
                "page_key": page_key,
                "page_hash": page_hash,
            }
        )

    if not shard_pages:
        raise HTTPException(status_code=400, detail="벡터화할 유효한 데이터가 없습니다.")

    rag_manager = get_rag_manager()

    result = await asyncio.to_thread(
        rag_manager.build_shard,
        shard_pages,
        form_type=None,
    )
    if not result:
        raise HTTPException(status_code=500, detail="Shard 생성에 실패했습니다.")

    shard_identifier, _shard_id = result

    merged = await asyncio.to_thread(rag_manager.merge_shard, shard_identifier)
    if not merged:
        raise HTTPException(
            status_code=500,
            detail="Shard를 base 인덱스에 병합하는 데 실패했습니다.",
        )

    await asyncio.to_thread(rag_manager.reload_index)
    total_vectors = rag_manager.count_examples()

    def _get_per_form(database):
        with database.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT form_type, SUM(vector_count) AS total_vectors FROM rag_vector_index GROUP BY form_type ORDER BY form_type")
            return [{"form_type": row[0], "vector_count": int(row[1] or 0)} for row in cursor.fetchall()]
    db = get_db()
    per_form = await db.run_sync(_get_per_form, db)
    return {
        "success": True,
        "message": "선택된 페이지들로부터 벡터 DB를 생성했습니다.",
        "processed_pages": len(shard_pages),
        "total_vectors": int(total_vectors),
        "per_form_type": per_form,
    }


RETAIL_USER_CSV_PATH = get_project_root() / "database" / "csv" / "retail_user.csv"


@router.get("/retail-user-csv")
async def get_retail_user_csv(
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """retail_user.csv をそのまま読み、一覧を返す（管理者専用）。担当・スーパータブで CSV 内容を表示。"""
    _ensure_admin(current_user)
    rows: List[Dict[str, str]] = []
    if not RETAIL_USER_CSV_PATH.exists():
        return {"rows": rows}
    try:
        with open(RETAIL_USER_CSV_PATH, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append({
                    "super_code": (row.get("소매처코드") or "").strip(),
                    "super_name": (row.get("소매처명") or "").strip(),
                    "person_id": (row.get("담당자ID") or "").strip(),
                    "person_name": (row.get("담당자명") or "").strip(),
                    "username": (row.get("ID") or "").strip(),
                })
    except Exception as e:
        logger.exception("retail_user.csv read failed: %s", e)
        raise HTTPException(status_code=500, detail="CSVの読み込みに失敗しました")
    return {"rows": rows}


class RetailUserCsvPutBody(BaseModel):
    """retail_user.csv 全体を上書きする用。"""
    rows: List[Dict[str, str]]


@router.put("/retail-user-csv")
async def put_retail_user_csv(
    body: RetailUserCsvPutBody,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """retail_user.csv を指定内容で上書き（管理者専用）。"""
    _ensure_admin(current_user)
    try:
        with open(RETAIL_USER_CSV_PATH, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["소매처코드", "소매처명", "담당자ID", "담당자명", "ID"])
            for r in body.rows:
                writer.writerow([
                    (r.get("super_code") or "").strip(),
                    (r.get("super_name") or "").strip(),
                    (r.get("person_id") or "").strip(),
                    (r.get("person_name") or "").strip(),
                    (r.get("username") or "").strip(),
                ])
    except Exception as e:
        logger.exception("retail_user.csv write failed: %s", e)
        raise HTTPException(status_code=500, detail="CSVの書き込みに失敗しました")
    return {"message": "保存しました", "rows_count": len(body.rows)}


DIST_RETAIL_CSV_PATH = get_project_root() / "database" / "csv" / "dist_retail.csv"


@router.get("/dist-retail-csv")
async def get_dist_retail_csv(
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """dist_retail.csv をそのまま読み、一覧を返す（管理者専用）。"""
    _ensure_admin(current_user)
    rows: List[Dict[str, str]] = []
    if not DIST_RETAIL_CSV_PATH.exists():
        return {"rows": rows}
    try:
        with open(DIST_RETAIL_CSV_PATH, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append({
                    "dist_code": (row.get("판매처코드") or "").strip(),
                    "dist_name": (row.get("판매처명") or "").strip(),
                    "super_code": (row.get("소매처코드") or "").strip(),
                    "super_name": (row.get("소매처명") or "").strip(),
                    "person_id": (row.get("담당자ID") or "").strip(),
                    "person_name": (row.get("담당자명") or "").strip(),
                })
    except Exception as e:
        logger.exception("dist_retail.csv read failed: %s", e)
        raise HTTPException(status_code=500, detail="CSVの読み込みに失敗しました")
    return {"rows": rows}


class DistRetailCsvPutBody(BaseModel):
    """dist_retail.csv 全体を上書きする用。"""
    rows: List[Dict[str, str]]


@router.put("/dist-retail-csv")
async def put_dist_retail_csv(
    body: DistRetailCsvPutBody,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """dist_retail.csv を指定内容で上書き（管理者専用）。"""
    _ensure_admin(current_user)
    try:
        with open(DIST_RETAIL_CSV_PATH, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["판매처코드", "판매처명", "소매처코드", "소매처명", "담당자ID", "담당자명"])
            for r in body.rows:
                writer.writerow([
                    (r.get("dist_code") or "").strip(),
                    (r.get("dist_name") or "").strip(),
                    (r.get("super_code") or "").strip(),
                    (r.get("super_name") or "").strip(),
                    (r.get("person_id") or "").strip(),
                    (r.get("person_name") or "").strip(),
                ])
    except Exception as e:
        logger.exception("dist_retail.csv write failed: %s", e)
        raise HTTPException(status_code=500, detail="CSVの書き込みに失敗しました")
    return {"message": "保存しました", "rows_count": len(body.rows)}


# ---- database/csv 汎用: 一覧・取得・上書き・アップロード・Excel ダウンロード ----
CSV_DIR = get_project_root() / "database" / "csv"
# ファイル名は英数字・アンダースコア・ハイフンのみ許可（パストラバーサル防止）
CSV_FILENAME_PATTERN = re.compile(r"^[a-zA-Z0-9_.-]+$")


def _resolve_csv_path(filename: str) -> Path:
    """filename（拡張子なし可）から database/csv 内の Path を返す。不正な filename は 400。"""
    base = filename.strip().lower().removesuffix(".csv")
    if not base or not CSV_FILENAME_PATTERN.match(base):
        raise HTTPException(status_code=400, detail="invalid filename")
    return CSV_DIR / f"{base}.csv"


@router.get("/csv-list")
async def get_csv_list(
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """database/csv 内の .csv ファイル名一覧（拡張子なし）を返す。"""
    _ensure_admin(current_user)
    exists = CSV_DIR.exists()
    logger.info("csv-list: CSV_DIR=%s exists=%s", str(CSV_DIR), exists)
    if not exists:
        return {"files": []}
    names = sorted(
        p.stem for p in CSV_DIR.iterdir() if p.suffix.lower() == ".csv" and p.is_file()
    )
    logger.info("csv-list: found %d files: %s", len(names), names)
    return {"files": names}


@router.get("/csv/{filename}")
async def get_csv(
    filename: str,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """CSV をそのまま読み、headers と rows（元の列名キー）を返す。"""
    _ensure_admin(current_user)
    path = _resolve_csv_path(filename)
    if not path.exists():
        return {"headers": [], "rows": []}
    try:
        with open(path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            rows = [dict(row) for row in reader]
        return {"headers": list(headers), "rows": rows}
    except Exception as e:
        logger.exception("csv read failed: %s", e)
        raise HTTPException(status_code=500, detail="CSVの読み込みに失敗しました")


class CsvPutBody(BaseModel):
    headers: List[str]
    rows: List[Dict[str, str]]


@router.put("/csv/{filename}")
async def put_csv(
    filename: str,
    body: CsvPutBody,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """CSV を指定内容で上書き。"""
    _ensure_admin(current_user)
    path = _resolve_csv_path(filename)
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(body.headers)
            for r in body.rows:
                row = [str(r.get(h, "")).strip() for h in body.headers]
                writer.writerow(row)
    except Exception as e:
        logger.exception("csv write failed: %s", e)
        raise HTTPException(status_code=500, detail="CSVの書き込みに失敗しました")
    return {"message": "保存しました", "rows_count": len(body.rows)}


@router.post("/csv/{filename}/upload")
async def upload_csv(
    filename: str,
    file: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """CSV または Excel をアップロードし、database/csv/{filename}.csv を上書き。"""
    _ensure_admin(current_user)
    path = _resolve_csv_path(filename)
    path.parent.mkdir(parents=True, exist_ok=True)
    suffix = (file.filename or "").lower().split(".")[-1]
    try:
        content = await file.read()
        if suffix == "xlsx":
            import pandas as pd
            df = pd.read_excel(io.BytesIO(content), engine="openpyxl")
            df = df.astype(str).fillna("")
            df.to_csv(path, index=False, encoding="utf-8", lineterminator="\n")
        else:
            # csv or default
            with open(path, "wb") as f:
                f.write(content)
    except Exception as e:
        logger.exception("csv upload failed: %s", e)
        raise HTTPException(status_code=500, detail="アップロードの処理に失敗しました")
    return {"message": "上書きしました", "path": str(path)}


@router.get("/csv/{filename}/download")
async def download_csv_excel(
    filename: str,
    format: str = Query("xlsx", alias="format"),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """CSV を Excel 形式でダウンロード。"""
    _ensure_admin(current_user)
    path = _resolve_csv_path(filename)
    if not path.exists():
        raise HTTPException(status_code=404, detail="ファイルがありません")
    if format.lower() != "xlsx":
        raise HTTPException(status_code=400, detail="format=xlsx のみ対応")
    try:
        import pandas as pd
        df = pd.read_csv(path, encoding="utf-8")
        buf = io.BytesIO()
        df.to_excel(buf, index=False, engine="openpyxl")
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={path.stem}.xlsx"},
        )
    except Exception as e:
        logger.exception("csv download xlsx failed: %s", e)
        raise HTTPException(status_code=500, detail="Excel の生成に失敗しました")


# ---- 판매처-소매처 RAG 정답지: created_by_user_id IS NOT NULL 문서의 item 中 得意先 / 受注先コード / 小売先コード ----
@router.get("/retail-rag-answer-items")
async def get_retail_rag_answer_items(
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    documents_current.created_by_user_id IS NOT NULL 인 문서에 속한 item만 조회.
    item_data에서 得意先, 受注先コード, 小売先コード 를 꺼내 중복 제거 후 반환 (RAG 정답지 후보).
    """
    _ensure_admin(current_user)
    def _fetch_retail_rag_items_sync(database):
        with database.get_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            cursor.execute("""
                SELECT DISTINCT i.item_data->>'得意先' AS "得意先", i.item_data->>'受注先コード' AS "受注先コード", i.item_data->>'小売先コード' AS "小売先コード"
                FROM items_current i INNER JOIN documents_current d ON d.pdf_filename = i.pdf_filename
                WHERE d.created_by_user_id IS NOT NULL ORDER BY "得意先", "受注先コード", "小売先コード"
            """)
            rows = cursor.fetchall()
        return [{"得意先": (r.get("得意先") or "").strip(), "受注先コード": (r.get("受注先コード") or "").strip(), "小売先コード": (r.get("小売先コード") or "").strip()} for r in rows]
    try:
        db = get_db()
        items = await db.run_sync(_fetch_retail_rag_items_sync, db)
        return {"items": items}
    except Exception as e:
        logger.exception("retail-rag-answer-items failed: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/retail-rag-answer-index/rebuild")
async def rebuild_retail_rag_answer_index(
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """판매처-소매처 RAG 정답지(得意先→受注先コード/小売先コード) 벡터 인덱스를 재구축해 DB에 저장."""
    _ensure_admin(current_user)
    try:
        rag = get_rag_manager()
        n = rag.build_retail_rag_answer_index()
        return {"message": "OK", "vector_count": n}
    except Exception as e:
        logger.exception("retail-rag-answer-index rebuild failed: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


# ---- 제품 RAG 정답지: created_by_user_id IS NOT NULL 문서의 item 中 商品名 / 商品コード / 仕切 / 本部長 ----
@router.get("/product-rag-answer-items")
async def get_product_rag_answer_items(
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    documents_current.created_by_user_id IS NOT NULL 인 문서에 속한 item만 조회.
    item_data에서 商品名, 商品コード, 仕切, 本部長 를 꺼내 중복 제거 후 반환 (제품 RAG 정답지 후보).
    """
    _ensure_admin(current_user)
    def _fetch_product_rag_items_sync(database):
        with database.get_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            cursor.execute("""
                SELECT DISTINCT
                    i.item_data->>'商品名' AS "商品名",
                    i.item_data->>'商品コード' AS "商品コード",
                    i.item_data->>'仕切' AS "仕切",
                    i.item_data->>'本部長' AS "本部長"
                FROM items_current i
                INNER JOIN documents_current d ON d.pdf_filename = i.pdf_filename
                WHERE d.created_by_user_id IS NOT NULL
                AND (i.item_data->>'商品名') IS NOT NULL AND (i.item_data->>'商品名') != ''
                AND (i.item_data->>'商品コード') IS NOT NULL AND (i.item_data->>'商品コード') != ''
                ORDER BY "商品名", "商品コード"
            """)
            rows = cursor.fetchall()
        return [
            {
                "商品名": (r.get("商品名") or "").strip(),
                "商品コード": (r.get("商品コード") or "").strip(),
                "仕切": (r.get("仕切") or "").strip(),
                "本部長": (r.get("本部長") or "").strip(),
            }
            for r in rows
        ]
    try:
        db = get_db()
        items = await db.run_sync(_fetch_product_rag_items_sync, db)
        return {"items": items}
    except Exception as e:
        logger.exception("product-rag-answer-items failed: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/product-rag-answer-index/rebuild")
async def rebuild_product_rag_answer_index(
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """제품 RAG 정답지(商品名→商品コード/仕切/本部長) 벡터 인덱스를 재구축해 DB에 저장."""
    _ensure_admin(current_user)
    try:
        rag = get_rag_manager()
        n = rag.build_product_rag_answer_index()
        return {"message": "OK", "vector_count": n}
    except Exception as e:
        logger.exception("product-rag-answer-index rebuild failed: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


# ---- 基準管理 (master_code.xlsx) ----
MASTER_CODE_PATH = get_project_root() / "master_code.xlsx"
MASTER_CODE_KEYS = ("a", "b", "c", "d", "e", "f")


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _read_master_code_xlsx() -> Tuple[List[str], List[Dict[str, str]]]:
    """master_code.xlsx を読み、1行目をヘッダー、2行目以降を rows (a,b,c,d,e,f) で返す。"""
    headers = ["取引先コード", "取引先名称", "代表スーパーコード", "代表スーパー名称", "担当ID", "担当者"]
    rows: List[Dict[str, str]] = []
    if not MASTER_CODE_PATH.exists():
        return headers, rows
    try:
        import openpyxl
        wb = openpyxl.load_workbook(MASTER_CODE_PATH, read_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(min_row=1, max_col=6, values_only=True))
        wb.close()
        if not all_rows:
            return headers, rows
        first = all_rows[0]
        headers = [_cell_str(first[i]) if i < len(first) else "" for i in range(6)]
        for row in all_rows[1:]:
            a = _cell_str(row[0]) if len(row) > 0 else ""
            b = _cell_str(row[1]) if len(row) > 1 else ""
            c = _cell_str(row[2]) if len(row) > 2 else ""
            d = _cell_str(row[3]) if len(row) > 3 else ""
            e = _cell_str(row[4]) if len(row) > 4 else ""
            f = _cell_str(row[5]) if len(row) > 5 else ""
            rows.append({"a": a, "b": b, "c": c, "d": d, "e": e, "f": f})
    except Exception:
        pass
    return headers, rows


class MasterCodeSaveRequest(BaseModel):
    headers: Optional[List[str]] = None
    rows: List[Dict[str, str]]


@router.get("/master-code")
async def get_master_code(
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """master_code.xlsx を読み、ヘッダーと全行を返す（管理者専用）。"""
    _ensure_admin(current_user)
    headers, rows = _read_master_code_xlsx()
    return {"headers": headers, "rows": rows}


@router.put("/master-code")
async def save_master_code(
    request: MasterCodeSaveRequest,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """master_code.xlsx に rows を書き戻す（管理者専用）。"""
    _ensure_admin(current_user)
    headers, _ = _read_master_code_xlsx()
    if request.headers and len(request.headers) >= 6:
        headers = [str(h) for h in request.headers[:6]]
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for r in request.rows:
            row = [r.get(k, "") for k in MASTER_CODE_KEYS]
            ws.append(row)
        wb.save(MASTER_CODE_PATH)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"保存に失敗しました: {e}")
    return {"success": True, "message": "保存しました。"}

