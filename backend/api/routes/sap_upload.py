"""
SAP 업로드 엑셀 파일 생성 API
- 데이터 입력/수식은 static/sap_upload_formulas.json 설정을 읽어 적용 (단일 소스)
"""
from pathlib import Path
from typing import List, Dict, Any, Optional
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from io import BytesIO
import csv
import re
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Alignment
import json

from database.registry import get_db
from backend.core.auth import get_current_user_optional
from backend.core.activity_log import log as activity_log

router = APIRouter()


def _get_project_root() -> Path:
    return Path(__file__).resolve().parent.parent.parent.parent


_RETAIL_USER_CSV = _get_project_root() / "database" / "csv" / "retail_user.csv"
_SAP_RETAIL_CSV = _get_project_root() / "database" / "csv" / "sap_retail.csv"
_SAP_PRODUCT_CSV = _get_project_root() / "database" / "csv" / "sap_product.csv"
_UNIT_PRICE_CSV = _get_project_root() / "database" / "csv" / "unit_price.csv"

# 소매처코드 → 담당자명 캐시 (D열). retail_user.csv: 소매처코드, 소매처명, 담당자ID, 담당자명, ID
_retail_code_to_담당자명: Optional[Dict[str, str]] = None
# 판매처코드 → 판매처명 (B열). sap_retail: C열 受注先コード=판매처코드 → 판매처명
_판매처코드_to_판매처명: Optional[Dict[str, str]] = None
# 소매처코드 → 소매처명 (K열). sap_retail: J열 小売先コード → 소매처명
_소매처코드_to_소매처명: Optional[Dict[str, str]] = None
# 제품코드 → 제품명 (L열). sap_product: M열 商品コード → 제품명
_제품코드_to_제품명: Optional[Dict[str, str]] = None
# 제품코드 → (2합환산값, 단일상자환산값) (N,O열). unit_price.csv
_제품코드_to_단가: Optional[Dict[str, tuple]] = None


def _get_담당자명_by_小売先コード(retail_code: str) -> str:
    """retail_user.csv에서 소매처코드=小売先コード인 행의 담당자명 반환 (없으면 '')."""
    global _retail_code_to_담당자명
    code = (retail_code or "").strip()
    if not code:
        return ""
    if _retail_code_to_담당자명 is None:
        _retail_code_to_담당자명 = {}
        if _RETAIL_USER_CSV.exists():
            try:
                with open(_RETAIL_USER_CSV, "r", encoding="utf-8") as f:
                    for row in csv.DictReader(f):
                        rc = (row.get("소매처코드") or "").strip()
                        name = (row.get("담당자명") or "").strip()
                        if rc and rc not in _retail_code_to_담당자명:
                            _retail_code_to_담당자명[rc] = name
            except Exception as e:
                print(f"❌ [retail_user] 읽기 오류: {e}")
    return (_retail_code_to_담당자명 or {}).get(code, "")


def _get_판매처명_by_受注先コード(code: str) -> str:
    """B열: C열 受注先コード를 sap_retail의 판매처코드와 매핑 → 판매처명."""
    global _판매처코드_to_판매처명
    code = (code or "").strip()
    if not code:
        return ""
    if _판매처코드_to_판매처명 is None:
        _판매처코드_to_판매처명 = {}
        if _SAP_RETAIL_CSV.exists():
            try:
                with open(_SAP_RETAIL_CSV, "r", encoding="utf-8") as f:
                    for row in csv.DictReader(f):
                        pc = (row.get("판매처코드") or "").strip()
                        pn = (row.get("판매처명") or "").strip()
                        if pc and pc not in _판매처코드_to_판매처명:
                            _판매처코드_to_판매처명[pc] = pn
            except Exception as e:
                print(f"❌ [sap_retail] 읽기 오류: {e}")
    return (_판매처코드_to_판매처명 or {}).get(code, "")


def _get_소매처명_by_小売先コード(code: str) -> str:
    """K열: J열 小売先コード를 sap_retail의 소매처코드와 매핑 → 소매처명."""
    global _소매처코드_to_소매처명
    code = (code or "").strip()
    if not code:
        return ""
    if _소매처코드_to_소매처명 is None:
        _소매처코드_to_소매처명 = {}
        if _SAP_RETAIL_CSV.exists():
            try:
                with open(_SAP_RETAIL_CSV, "r", encoding="utf-8") as f:
                    for row in csv.DictReader(f):
                        rc = (row.get("소매처코드") or "").strip()
                        rn = (row.get("소매처명") or "").strip()
                        if rc and rc not in _소매처코드_to_소매처명:
                            _소매처코드_to_소매처명[rc] = rn
            except Exception as e:
                print(f"❌ [sap_retail] 읽기 오류: {e}")
    return (_소매처코드_to_소매처명 or {}).get(code, "")


def _get_제품명_by_商品コード(code: str) -> str:
    """L열: M열 商品コード를 sap_product의 제품코드와 매핑 → 제품명."""
    global _제품코드_to_제품명
    code = (code or "").strip()
    if not code:
        return ""
    if _제품코드_to_제품명 is None:
        _제품코드_to_제품명 = {}
        if _SAP_PRODUCT_CSV.exists():
            try:
                with open(_SAP_PRODUCT_CSV, "r", encoding="utf-8") as f:
                    for row in csv.DictReader(f):
                        c = (row.get("제품코드") or "").strip()
                        n = (row.get("제품명") or "").strip()
                        if c and c not in _제품코드_to_제품명:
                            _제품코드_to_제품명[c] = n
            except Exception as e:
                print(f"❌ [sap_product] 읽기 오류: {e}")
    return (_제품코드_to_제품명 or {}).get(code, "")


def _get_unit_price_by_商品コード(code: str) -> tuple:
    """N,O열: M열 商品コード를 unit_price의 제품코드와 매핑 → (2합환산값, 단일상자환산값)."""
    global _제품코드_to_단가
    code = (code or "").strip()
    if not code:
        return ("", "")
    if _제품코드_to_단가 is None:
        _제품코드_to_단가 = {}
        if _UNIT_PRICE_CSV.exists():
            try:
                with open(_UNIT_PRICE_CSV, "r", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    reader.fieldnames = [c.strip().lstrip("\ufeff") for c in (reader.fieldnames or [])]
                    for row in reader:
                        c = (row.get("제품코드") or "").strip()
                        n = (row.get("2합환산값") or "").strip()
                        o = (row.get("단일상자환산값") or "").strip()
                        if c and c not in _제품코드_to_단가:
                            _제품코드_to_단가[c] = (n, o)
            except Exception as e:
                print(f"❌ [unit_price] 읽기 오류: {e}")
    t = (_제품코드_to_단가 or {}).get(code, ("", ""))
    return t if isinstance(t, tuple) else ("", "")


# ---------- 설정 기반 규칙 해석 (rule → 값) ----------
_FIELD_FALLBACK = {
    "得意先名": "得意先",
    "ケース入数": "入数",
    "最終金額": "金額",
    "最終請求金額": "請求金額",
    "最終請求額": "請求額",
}


def _safe_eval_expr(expr: str, item_data: Dict[str, Any]) -> Any:
    """
    item_data 필드명만 사용하는 수식 안전 계산.
    예: "条件+条件小数部*0.01" → item_data["条件"] + item_data["条件小数部"]*0.01
    """
    if not expr or not expr.strip():
        return ""
    expr = expr.strip()
    tokens = re.findall(r"[^\d\s\+\-\*\/\(\)\.\,\=\>\<\'\"]+", expr)
    locals_map = {}
    for t in set(tokens):
        if t in ("and", "or", "not", "True", "False"):
            continue
        v = item_data.get(t)
        if (v is None or v == "") and t in _FIELD_FALLBACK:
            v = item_data.get(_FIELD_FALLBACK[t])
        if v is None or v == "":
            v = 0
        try:
            if isinstance(v, str):
                v = v.replace(",", "").replace("，", "").strip()
            v = float(v)
        except (TypeError, ValueError):
            v = str(v) if v else ""
        locals_map[t] = v
    try:
        work = expr
        for k, v in sorted(locals_map.items(), key=lambda x: -len(x[0])):
            if isinstance(v, str):
                work = work.replace(k, repr(v))
            else:
                work = work.replace(k, str(v))
        return eval(work)
    except Exception:
        return ""


def _apply_data_rule(rule_spec: Any, item_data: Dict[str, Any]) -> Any:
    """
    설정의 rule 한 건을 item_data에 적용해 값 반환.
    rule_spec: 문자열(필드명) | {"field": "필드명"} | {"cond": [...]} | {"expr": "수식"}
    """
    if rule_spec is None:
        return ""
    if isinstance(rule_spec, str):
        return item_data.get(rule_spec, "") or ""
    if not isinstance(rule_spec, dict):
        return ""
    # {"field": "필드명"}
    if "field" in rule_spec:
        return item_data.get(rule_spec["field"], "") or ""
    # {"field_digits": "필드명"} → 필드값에서 숫자만 추출
    if "field_digits" in rule_spec:
        val = item_data.get(rule_spec["field_digits"], "") or ""
        if isinstance(val, str):
            numbers = re.findall(r"\d+", val.replace(",", ""))
            return "".join(numbers) if numbers else ""
        return str(val) if val else ""
    # {"cond": [ {"if_field":"A", "if_eq":"個", "then_field":"B"} | {"then_expr":"..."} }, ...]}
    if "cond" in rule_spec:
        for c in rule_spec["cond"]:
            if_field = c.get("if_field")
            if_eq = c.get("if_eq")
            if if_field is None:
                continue
            if str(item_data.get(if_field, "")) != str(if_eq):
                continue
            if "then_field" in c:
                return item_data.get(c["then_field"], "") or ""
            if "then_expr" in c:
                return _safe_eval_expr(c["then_expr"], item_data)
        return ""
    # {"expr": "수식"}
    if "expr" in rule_spec:
        return _safe_eval_expr(rule_spec["expr"], item_data)
    return ""


def _get_rule_from_by_form(by_form_value: Any) -> Any:
    """byForm[form] 값에서 실행용 rule만 추출 (문자열이면 field 규칙으로 취급)"""
    if by_form_value is None:
        return None
    if isinstance(by_form_value, str):
        return {"field": by_form_value} if by_form_value.strip() else None
    if isinstance(by_form_value, dict) and "rule" in by_form_value:
        return by_form_value["rule"]
    if isinstance(by_form_value, dict) and (
        "field" in by_form_value or "field_digits" in by_form_value
        or "cond" in by_form_value or "expr" in by_form_value
    ):
        return by_form_value
    return None


def get_all_items_current(
    db,
    data_year: Optional[int] = None,
    data_month: Optional[int] = None,
) -> List[Dict[str, Any]]:
    """
    items_current에서 아이템 조회.
    조건: page_role='detail', documents_current.created_by_user_id IS NOT NULL,
    (선택) data_year/data_month 일치.
    Returns:
        아이템 리스트 (pdf_filename, page_number, item_data, form_type 등)
    """
    try:
        with db.get_connection() as conn:
            from psycopg2.extras import RealDictCursor
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            ym_clause = ""
            params: tuple = ()
            if data_year is not None and data_month is not None:
                ym_clause = " AND d.data_year = %s AND d.data_month = %s"
                params = (data_year, data_month)
            cursor.execute(
                """
                SELECT 
                    i.item_id,
                    i.pdf_filename,
                    i.page_number,
                    i.item_order,
                    i.item_data::text as item_data,
                    d.form_type
                FROM items_current i
                INNER JOIN documents_current d 
                    ON i.pdf_filename = d.pdf_filename
                   AND d.created_by_user_id IS NOT NULL
                JOIN page_data_current p
                    ON i.pdf_filename = p.pdf_filename
                   AND i.page_number = p.page_number
                WHERE p.page_role = 'detail'
                  AND i.first_review_checked = TRUE
                  AND i.second_review_checked = TRUE
                """ + ym_clause + """
                ORDER BY i.pdf_filename, i.page_number, i.item_order
            """,
                params,
            )
            rows = cursor.fetchall()
            result = []
            for row in rows:
                item_dict = dict(row)
                # item_data가 문자열이면 JSON 파싱
                if isinstance(item_dict.get('item_data'), str):
                    try:
                        item_dict['item_data'] = json.loads(item_dict['item_data'])
                    except:
                        item_dict['item_data'] = {}
                result.append(item_dict)
            return result
    except Exception as e:
        print(f"❌ [get_all_items_current] 오류: {e}")
        import traceback
        traceback.print_exc()
        raise


def _load_formulas_config() -> Dict[str, Any]:
    """static/sap_upload_formulas.json 로드 (없으면 기본값)."""
    path = _get_formulas_path()
    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"❌ [formulas config] 읽기 오류: {e}")
    return _default_formulas()


def _strip_numeric_unit(val: Any) -> str:
    """
    T열 등 연산에 쓰는 값에서 '48個', '1,234円' 같은 단위 제거.
    반환: 숫자만 있는 문자열 (빈 문자열 또는 '48', '1234.5' 등)
    """
    if val is None or val == "":
        return ""
    if isinstance(val, (int, float)):
        return str(val)
    s = str(val).strip()
    s = re.sub(r"[,，\s]", "", s)  # 천단위 콤마·공백 제거
    m = re.match(r"^[\d.]+", s)
    return m.group(0) if m else ""


def process_item_for_sap(
    item: Dict[str, Any],
    form_type: Optional[str],
    config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """
    아이템 데이터를 SAP 양식에 맞게 가공.
    config(static/sap_upload_formulas.json)의 dataInputColumns 규칙을 적용.
    """
    item_data = item.get('item_data', {})
    if isinstance(item_data, str):
        try:
            item_data = json.loads(item_data)
        except Exception:
            item_data = {}
    if not isinstance(item_data, dict):
        item_data = {}

    form_type = (form_type or "01").strip()
    if form_type not in ("01", "02", "03", "04", "05"):
        form_type = "01"

    if config is None:
        config = _load_formulas_config()

    data_columns = config.get("dataInputColumns") or []
    result = {}
    for col_spec in data_columns:
        col = col_spec.get("column")
        if not col:
            continue
        by_form = col_spec.get("byForm") or {}
        raw = by_form.get(form_type)
        rule = _get_rule_from_by_form(raw)
        if rule is None:
            result[col] = ""
            continue
        val = _apply_data_rule(rule, item_data)
        if val is None:
            val = ""
        result[col] = val

    # T열은 U열 수식(P×N+R×O+T)에 사용되므로 단위(個, 円 등) 제거
    if result.get("T") not in (None, ""):
        result["T"] = _strip_numeric_unit(result["T"])

    return result


def _safe_float(val: Any) -> float:
    """null/빈 문자열이면 0, 아니면 float 변환 (변환 실패 시 0)."""
    if val is None or val == "":
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).strip()
        return float(s) if s else 0.0
    except (ValueError, TypeError):
        return 0.0


def _compute_u_row(processed: Dict[str, Any]) -> float:
    """U열(個数計) = P×N + R×O + T. null/빈 값은 0으로 계산."""
    p = _safe_float(processed.get("P"))
    n = _safe_float(processed.get("N"))
    r = _safe_float(processed.get("R"))
    o = _safe_float(processed.get("O"))
    t = _safe_float(processed.get("T"))
    return p * n + r * o + t


def get_column_letters_a_to_bb() -> List[str]:
    """
    A~BB열까지 모든 열 문자 리스트 생성
    
    Returns:
        ['A', 'B', ..., 'Z', 'AA', 'AB', ..., 'AZ', 'BA', 'BB']
    """
    columns = []
    # A~BB까지 총 54개 열 (A=1, Z=26, AA=27, AZ=52, BA=53, BB=54)
    for i in range(1, 55):  # 1부터 54까지
        columns.append(get_column_letter(i))
    return columns


def create_sap_excel(
    items: List[Dict[str, Any]],
    template_path: Optional[str] = None,
    data_year: Optional[int] = None,
    data_month: Optional[int] = None,
) -> BytesIO:
    """
    SAP 업로드용 엑셀 파일 생성.
    data_year/data_month 있으면 W열(発生月)에 "YYYY.MM" 형식으로 채움.
    """
    from pathlib import Path
    
    # 템플릿 파일이 있으면 로드, 없으면 새로 생성
    if template_path and Path(template_path).exists():
        wb = load_workbook(template_path)
        ws = wb.active
        # 첫 행(1행)은 컬럼명이므로 유지
        # 기존 데이터 행 삭제 (2행부터 삭제)
        if ws.max_row >= 2:
            ws.delete_rows(2, ws.max_row - 1)
        data_start_row = 2  # 2행부터 데이터 시작
    else:
        # 새 워크북 생성 (템플릿이 없으면 컬럼명도 없으므로 경고)
        # 실제로는 템플릿 파일을 사용하는 것이 권장됨
        wb = Workbook()
        ws = wb.active
        ws.title = "SAP Upload"
        # 템플릿이 없으면 첫 행에 컬럼명이 없으므로
        # 2행부터 데이터 시작 (1행은 비어있음)
        data_start_row = 2
        # A~BB열까지 모든 열이 존재하도록 보장 (최소한 한 행이라도 있어야 열이 생성됨)
        # 첫 번째 데이터 행을 미리 생성하여 모든 열이 존재하도록 함
        all_columns = get_column_letters_a_to_bb()
        for col_letter in all_columns:
            ws[f'{col_letter}{data_start_row}'] = ''
    
    config = _load_formulas_config()
    data_columns = [c["column"] for c in (config.get("dataInputColumns") or [])]
    formula_columns = config.get("excelFormulaColumns") or []

    # 수식 템플릿에서 행 번호 치환 (3 → 실제 row_num). 예: =P3*N3 → =P{row}*N{row}
    def formula_for_row(formula_tpl: str, row_num: int) -> str:
        if not formula_tpl or not str(formula_tpl).strip().startswith("="):
            return formula_tpl
        return re.sub(r"([A-Z]+)\d+", lambda m: m.group(1) + str(row_num), formula_tpl)

    row_num = data_start_row
    for item in items:
        form_type = item.get("form_type", "01")
        processed = process_item_for_sap(item, form_type, config)
        # lookup 열: item_data 없이 CSV 매핑으로 채움
        item_data = item.get("item_data") or {}
        if isinstance(item_data, str):
            try:
                item_data = json.loads(item_data)
            except Exception:
                item_data = {}
        retail_cd = (processed.get("J") or item_data.get("小売先コード") or item_data.get("小売先CD") or "").strip()
        processed["D"] = _get_담당자명_by_小売先コード(retail_cd)
        processed["B"] = _get_판매처명_by_受注先コード(processed.get("C") or "")
        processed["K"] = _get_소매처명_by_小売先コード(processed.get("J") or "")
        processed["L"] = _get_제품명_by_商品コード(processed.get("M") or "")
        n_val, o_val = _get_unit_price_by_商品コード(processed.get("M") or "")
        processed["N"], processed["O"] = n_val, o_val
        if data_year is not None and data_month is not None:
            processed["W"] = f"{data_year}.{data_month:02d}"

        for col in data_columns:
            val = processed.get(col)
            if val is not None and val != "":
                try:
                    ws[f"{col}{row_num}"] = val
                except Exception:
                    pass
        # W열(発生月): 대상기간으로 무조건 기록 (data_columns 순서와 무관)
        if data_year is not None and data_month is not None:
            try:
                ws[f"W{row_num}"] = f"{data_year}.{data_month:02d}"
            except Exception:
                pass
        # U열(個数計): P×N + R×O + T (null/빈 값은 0으로 계산)
        try:
            ws[f"U{row_num}"] = _compute_u_row(processed)
        except Exception:
            pass

        for fc in formula_columns:
            col = fc.get("column")
            if col == "U":
                continue  # U는 위에서 계산값으로 이미 기록
            formula_tpl = fc.get("formula")
            if col and formula_tpl:
                ws[f"{col}{row_num}"] = formula_for_row(formula_tpl, row_num)

        row_num += 1
    
    # 파일을 메모리에 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def _get_sap_available_year_months_sync(db):
    with db.get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT DISTINCT d.data_year AS y, d.data_month AS m
            FROM documents_current d
            WHERE d.created_by_user_id IS NOT NULL
              AND d.data_year IS NOT NULL AND d.data_month IS NOT NULL
              AND EXISTS (SELECT 1 FROM page_data_current p WHERE p.pdf_filename = d.pdf_filename AND p.page_role = 'detail')
            ORDER BY y DESC, m DESC
        """)
        return [{"year": r[0], "month": r[1]} for r in cursor.fetchall()]


@router.get("/available-year-months")
async def get_sap_available_year_months(db=Depends(get_db)):
    """SAP 대상 문서가 있는 연월 목록."""
    try:
        year_months = await db.run_sync(_get_sap_available_year_months_sync, db)
        return {"year_months": year_months}
    except Exception as e:
        print(f"❌ [get_sap_available_year_months] 오류: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/check-document")
async def check_sap_document(
    db=Depends(get_db),
    pdf: Optional[str] = None,
):
    """
    특정 문서가 SAP 대상 목록에 안 나오는 이유 진단.
    pdf: 파일명 (예: 三菱食品東日本_2025.01 (2)-1-4.pdf)
    반환: in_db, created_by_user_id, data_year, data_month, has_detail_page, reason
    """
    if not pdf or not pdf.strip():
        return {"ok": False, "reason": "pdf 파라미터가 비어 있습니다."}
    pdf = pdf.strip()
    try:
        def _check_sap_doc_sync():
            from psycopg2.extras import RealDictCursor
            with db.get_connection() as conn:
                cursor = conn.cursor(cursor_factory=RealDictCursor)
                cursor.execute(
                    "SELECT pdf_filename, created_by_user_id, data_year, data_month, form_type FROM documents_current WHERE pdf_filename = %s",
                    (pdf,),
                )
                row = cursor.fetchone()
            if not row:
                return {"ok": True, "pdf_filename": pdf, "in_db": False, "created_by_user_id": None, "data_year": None, "data_month": None, "has_detail_page": False, "reason": "documents_current에 해당 pdf_filename이 없습니다. (아카이브에 있거나 미등록)"}
            d = dict(row)
            created_by = d.get("created_by_user_id")
            data_year = d.get("data_year")
            data_month = d.get("data_month")
            with db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT 1 FROM page_data_current WHERE pdf_filename = %s AND page_role = 'detail' LIMIT 1", (pdf,))
                has_detail = cursor.fetchone() is not None
            reasons = []
            if created_by is None:
                reasons.append("created_by_user_id가 NULL입니다. (업로드 시 로그인 사용자로 저장된 문서만 대상)")
            if data_year is None or data_month is None:
                reasons.append("data_year 또는 data_month가 NULL입니다. (연월이 지정된 문서만 대상)")
            if not has_detail:
                reasons.append("detail 페이지가 없습니다. (page_data_current에 page_role='detail'인 페이지가 없음)")
            reason = "; ".join(reasons) if reasons else "조건은 모두 만족합니다. 연월 선택이 이 문서의 data_year/data_month와 일치하는지 확인하세요."
            return {"ok": True, "pdf_filename": pdf, "in_db": True, "created_by_user_id": created_by, "data_year": data_year, "data_month": data_month, "has_detail_page": has_detail, "reason": reason}
        return await db.run_sync(_check_sap_doc_sync)
    except Exception as e:
        print(f"❌ [check_sap_document] 오류: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/documents")
async def get_sap_documents(
    db=Depends(get_db),
    year: Optional[int] = None,
    month: Optional[int] = None,
):
    """
    조건 충족 문서 목록을 양식지(form_type)별로 반환.
    조건: created_by_user_id IS NOT NULL, detail 페이지 있음, (선택) data_year/month 일치.
    2次検討まで完了した行のみカウント（SAP対象は2次検討済みのみ）。
    Returns:
        by_form: { "01": [{ pdf_filename, item_count }], ... },
        total_items: 전체 행 수（2次検討済みのみ）
    """
    if year is None or month is None:
        return {"by_form": {}, "total_items": 0}
    try:
        def _get_sap_docs_sync():
            from psycopg2.extras import RealDictCursor
            with db.get_connection() as conn:
                cursor = conn.cursor(cursor_factory=RealDictCursor)
                cursor.execute("""
                    SELECT d.pdf_filename, d.form_type, COUNT(i.item_id) AS item_count
                    FROM documents_current d
                    INNER JOIN page_data_current p ON p.pdf_filename = d.pdf_filename AND p.page_role = 'detail'
                    INNER JOIN items_current i ON i.pdf_filename = d.pdf_filename AND i.page_number = p.page_number
                        AND i.first_review_checked = TRUE AND i.second_review_checked = TRUE
                    WHERE d.created_by_user_id IS NOT NULL AND d.data_year = %s AND d.data_month = %s
                    GROUP BY d.pdf_filename, d.form_type ORDER BY d.form_type, d.pdf_filename
                """, (year, month))
                rows = cursor.fetchall()
            by_form: Dict[str, List[Dict[str, Any]]] = {}
            total_items = 0
            for r in rows:
                row = dict(r)
                ft = (row.get("form_type") or "01").strip()
                if ft not in by_form:
                    by_form[ft] = []
                by_form[ft].append({"pdf_filename": row.get("pdf_filename", ""), "item_count": int(row.get("item_count", 0))})
                total_items += int(row.get("item_count", 0))
            return {"by_form": by_form, "total_items": total_items}
        return await db.run_sync(_get_sap_docs_sync)
    except Exception as e:
        print(f"❌ [get_sap_documents] 오류: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/preview")
async def preview_sap_excel(
    db=Depends(get_db),
    year: Optional[int] = Query(None, description="対象年"),
    month: Optional[int] = Query(None, description="対象月"),
):
    """
    SAP 엑셀 파일 미리보기 (JSON). 연월 지정 시 해당 기간 문서만.
    """
    try:
        from pathlib import Path

        items = await db.run_sync(get_all_items_current, db, year, month)
        
        # 템플릿 파일에서 컬럼명 읽기
        project_root = Path(__file__).parent.parent.parent.parent
        template_path = project_root / "static" / "sap_upload.xlsx"
        
        column_names = []
        if template_path.exists():
            wb = load_workbook(template_path, read_only=True)
            ws = wb.active
            # 첫 행(1행)에서 모든 컬럼명 읽기
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col_idx).value
                if cell_value:
                    column_names.append(str(cell_value))
                else:
                    # 값이 없어도 열은 존재하므로 빈 문자열 추가
                    column_names.append('')
            wb.close()
        else:
            # 템플릿이 없으면 A~BB까지 열 문자로 표시
            all_columns = get_column_letters_a_to_bb()
            column_names = all_columns
        
        if not items:
            return {
                "total_items": 0,
                "preview_rows": [],
                "column_names": column_names,
                "message": "データがありません"
            }
        
        config = _load_formulas_config()
        preview_items = items[:50]
        preview_data = []

        for item in preview_items:
            form_type = item.get("form_type", "01")
            processed = process_item_for_sap(item, form_type, config)
            # lookup 열: D,B,K,L,N,O
            item_data = item.get("item_data") or {}
            if isinstance(item_data, str):
                try:
                    item_data = json.loads(item_data)
                except Exception:
                    item_data = {}
            retail_cd = (processed.get("J") or item_data.get("小売先コード") or item_data.get("小売先CD") or "").strip()
            processed["D"] = _get_담당자명_by_小売先コード(retail_cd)
            processed["B"] = _get_판매처명_by_受注先コード(processed.get("C") or "")
            processed["K"] = _get_소매처명_by_小売先コード(processed.get("J") or "")
            processed["L"] = _get_제품명_by_商品コード(processed.get("M") or "")
            n_val, o_val = _get_unit_price_by_商品コード(processed.get("M") or "")
            processed["N"], processed["O"] = n_val, o_val
            if year is not None and month is not None:
                processed["W"] = f"{year}.{month:02d}"
            processed["U"] = _compute_u_row(processed)

            row_data: Dict[str, Any] = {
                "pdf_filename": item.get("pdf_filename", ""),
                "page_number": item.get("page_number", 0),
                "form_type": form_type,
            }
            all_columns = get_column_letters_a_to_bb()
            for col_letter in all_columns:
                row_data[col_letter] = processed.get(col_letter, "")

            preview_data.append(row_data)
        
        return {
            "total_items": len(items),
            "preview_rows": preview_data,
            "column_names": column_names,
            "message": f"全{len(items)}件のデータがあります（最初の50件を表示）"
        }
    except Exception as e:
        print(f"❌ [preview_sap_excel] 오류: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/download")
async def download_sap_excel(
    db=Depends(get_db),
    current_user: dict = Depends(get_current_user_optional),
    year: Optional[int] = Query(None, description="対象年"),
    month: Optional[int] = Query(None, description="対象月"),
):
    """
    SAP 업로드용 엑셀 파일 다운로드. year/month 지정 시 해당 기간 문서의 item만 취합.
    """
    try:
        items = await db.run_sync(get_all_items_current, db, year, month)

        if not items:
            raise HTTPException(status_code=404, detail="データがありません")

        from pathlib import Path
        project_root = Path(__file__).parent.parent.parent.parent
        template_path = project_root / "static" / "sap_upload.xlsx"
        template_file = str(template_path) if template_path.exists() else None

        excel_file = create_sap_excel(items, template_file, data_year=year, data_month=month)

        from datetime import datetime
        ym = f"{year}{month:02d}_" if year and month else ""
        filename = f"SAP_Upload_{ym}{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        if current_user:
            activity_log(current_user.get("username"), "SAP 엑셀 다운로드")
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ [download_sap_excel] 오류: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/column-names")
async def get_sap_column_names():
    """
    SAP 업로드 템플릿(sap_upload.xlsx) 1행에서 컬럼명 목록 반환.
    산식 표시 시 실제 컬럼명 표시용.
    """
    from pathlib import Path
    project_root = Path(__file__).resolve().parent.parent.parent.parent
    template_path = project_root / "static" / "sap_upload.xlsx"
    column_names = []
    if template_path.exists():
        wb = load_workbook(template_path, read_only=True)
        ws = wb.active
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            if cell_value:
                column_names.append(str(cell_value))
            else:
                column_names.append("")
        wb.close()
    else:
        column_names = get_column_letters_a_to_bb()
    return {"column_names": column_names}


# ========== SAP 산식 설정 (양식지별 편집용) ==========
def _get_formulas_path() -> Path:
    project_root = Path(__file__).resolve().parent.parent.parent.parent
    return project_root / "static" / "sap_upload_formulas.json"


def _default_formulas() -> Dict[str, Any]:
    """
    sap_upload.md 기반 기본 산식 (파일 없을 때 fallback).
    byForm: 문자열=필드명, 객체=rule(field/cond/expr). 백엔드가 해석해 적용.
    B/K/L/N/O/W: lookup 또는 메타는 별도 처리·공란.
    """
    return {
        "dataInputColumns": [
            {"column": "B", "byForm": {"01": "", "02": "", "03": "", "04": "", "05": ""}},  # C열→sap_retail 판매처명 lookup
            {"column": "C", "byForm": {"01": {"field": "受注先コード"}, "02": {"field": "受注先コード"}, "03": {"field": "受注先コード"}, "04": {"field": "受注先コード"}, "05": {"field": "受注先コード"}}},
            {"column": "D", "byForm": {"01": {"field": "담당자명"}, "02": {"field": "담당자명"}, "03": {"field": "담당자명"}, "04": {"field": "담당자명"}, "05": {"field": "담당자명"}}},  # J열→retail_user 담당자명은 create_sap_excel에서 덮어씀
            {"column": "J", "byForm": {"01": {"field": "小売先コード"}, "02": {"field": "小売先コード"}, "03": {"field": "小売先コード"}, "04": {"field": "小売先コード"}, "05": {"field": "小売先コード"}}},
            {"column": "K", "byForm": {"01": "", "02": "", "03": "", "04": "", "05": ""}},  # C열→sap_retail 소매처명 lookup
            {"column": "L", "byForm": {"01": "", "02": "", "03": "", "04": "", "05": ""}},  # M열→sap_product 제품명 lookup
            {"column": "M", "byForm": {"01": {"field": "商品コード"}, "02": {"field": "商品コード"}, "03": {"field": "商品コード"}, "04": {"field": "商品コード"}, "05": {"field": "商品コード"}}},
            {"column": "N", "byForm": {"01": "", "02": "", "03": "", "04": "", "05": ""}},  # unit_price 2합환산 lookup
            {"column": "O", "byForm": {"01": "", "02": "", "03": "", "04": "", "05": ""}},  # unit_price 단일상자환산 lookup
            {"column": "P", "byForm": {"01": "", "02": "", "03": {"field": "ケース"}, "04": "", "05": ""}},
            {"column": "R", "byForm": {"01": "", "02": "", "03": {"field": "バラ"}, "04": "", "05": ""}},
            {
                "column": "T",
                "byForm": {
                    "01": {"cond": [{"if_field": "数量単位", "if_eq": "個", "then_field": "数量"}, {"if_field": "数量単位", "if_eq": "CS", "then_expr": "ケース入数*数量"}]},
                    "02": {"field": "取引数量合計"},
                    "03": {"field": "バラ"},
                    "04": {"field": "対象数量又は金額"},
                    "05": "",
                },
            },
            {"column": "W", "byForm": {"01": "", "02": "", "03": "", "04": "", "05": ""}},  # 연월 메타
            {
                "column": "AL",
                "byForm": {
                    "01": {"field": "最終金額"},
                    "02": {"field": "最終金額"},
                    "03": {"field": "最終請求金額"},
                    "04": {"field": "最終金額"},
                    "05": {"field": "最終請求額"},
                },
            },
        ],
        "excelFormulaColumns": [
            {"column": "U", "formula": "=P3*N3 + R3*O3 + T3", "description": "P×N + R×O + T"},
        ],
    }


@router.get("/formulas")
async def get_sap_formulas():
    """
    SAP 산식 설정 조회 (양식지 01~05별 데이터 입력, 엑셀 수식).
    파일 없으면 기본값 반환.
    """
    path = _get_formulas_path()
    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"❌ [get_sap_formulas] 읽기 오류: {e}")
    return _default_formulas()


@router.put("/formulas")
async def put_sap_formulas(body: Dict[str, Any]):
    """
    SAP 산식 설정 저장 (양식지별 편집 결과).
    """
    path = _get_formulas_path()
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(body, f, ensure_ascii=False, indent=2)
        return {"ok": True}
    except Exception as e:
        print(f"❌ [put_sap_formulas] 저장 오류: {e}")
        raise HTTPException(status_code=500, detail=str(e))
